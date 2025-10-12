from flask import Blueprint, request, jsonify, send_file, abort
from flask_restx import Resource
from marshmallow import ValidationError
from datetime import datetime, date
from sqlalchemy import and_, or_
from sqlalchemy.exc import IntegrityError, OperationalError, SQLAlchemyError
import logging
from .. import db
import os
import uuid
import json
from werkzeug.utils import secure_filename
from flask import request, current_app
from flask_restx import Resource, reqparse
from ..models import Transaction, AssetTransaction, Warehouse, FixedAsset, Branch, Category
from ..schemas import (
    TransactionSchema, TransactionCreateSchema, 
    AssetTransactionSchema, AssetTransactionCreateSchema
)
from flask_jwt_extended import get_jwt_identity, jwt_required, verify_jwt_in_request, decode_token, get_jwt
from flask_jwt_extended.exceptions import JWTExtendedException
from ..utils import check_permission, error_response, create_error_response, create_validation_error_response
from ..swagger import transactions_ns, asset_transactions_ns, add_standard_responses, api
from ..swagger_models import (
    transaction_model, transaction_input_model, transaction_create_model,
    asset_transaction_model, asset_transaction_input_model,
    pagination_model, error_model, success_model
)
from werkzeug.datastructures import FileStorage
from sqlalchemy.orm import noload
import pandas as pd
from io import BytesIO
from datetime import datetime, date
from flask import send_file
import openpyxl
import openpyxl.styles

bp = Blueprint("transactions", __name__, url_prefix="/transactions")

def authenticate_for_download():
    """
    Custom authentication function for download endpoints.
    Tries JWT header first, then token query parameter.
    Returns user_id if authenticated, None otherwise.
    """
    # Try header-based JWT first
    try:
        verify_jwt_in_request()
        return get_jwt_identity()
    except JWTExtendedException:
        pass
    
    # Try token query parameter
    token = request.args.get('token')
    if not token:
        return None
    
    try:
        decoded_token = decode_token(token)
        # Extract user identity from the decoded token
        return decoded_token.get('sub')  # 'sub' is the standard claim for user identity
    except JWTExtendedException:
        return None
    except Exception:
        return None

# Initialize schemas
transaction_schema = TransactionSchema()
transactions_schema = TransactionSchema(many=True)
transaction_create_schema = TransactionCreateSchema()
asset_transaction_schema = AssetTransactionSchema()
asset_transactions_schema = AssetTransactionSchema(many=True)
asset_transaction_create_schema = AssetTransactionCreateSchema()
file_upload_parser = reqparse.RequestParser()
file_upload_parser.add_argument('attached_file', 
                               location='files',
                               type=FileStorage,
                               required=False,
                               help='File attachment')
file_upload_parser.add_argument('data', 
                               location='form',
                               type=str,
                               required=False,
                               help='''Transaction data as JSON string. Example:
{
  "date": "2025-09-24",
  "description": "string", 
  "reference_number": "string",
  "warehouse_id": 0,
  "transaction_type": true,
  "asset_transactions": [
    {
      "asset_id": 0,
      "quantity": 1,
      "amount": 0
    }
  ]
}''')



@transactions_ns.route("/")
class TransactionList(Resource):
    @transactions_ns.doc('list_transactions', security='Bearer Auth')
    @transactions_ns.response(200, 'Successfully retrieved transactions', pagination_model)
    @transactions_ns.response(400, 'Bad Request', error_model)
    @transactions_ns.response(401, 'Unauthorized', error_model)
    @transactions_ns.response(403, 'Forbidden', error_model)
    @transactions_ns.response(500, 'Internal Server Error', error_model)
    @transactions_ns.response(503, 'Service Unavailable', error_model)
    @transactions_ns.param('page', 'Page number', type=int, default=1)
    @transactions_ns.param('per_page', 'Items per page', type=int, default=10)
    @transactions_ns.param('branch_id', 'Filter by branch ID', type=int)
    @transactions_ns.param('warehouse_id', 'Filter by warehouse ID', type=int)
    @transactions_ns.param('date_from', 'Filter transactions from date (YYYY-MM-DD)', type=str)
    @transactions_ns.param('date_to', 'Filter transactions to date (YYYY-MM-DD)', type=str)
    @transactions_ns.param('search', 'Search in description or reference number', type=str)
    @jwt_required()
    def get(self):
        """Get all transactions with pagination and filtering"""
        error = check_permission("can_make_transaction")  # Using transaction permission for now
        if error:
            return error

        page = request.args.get("page", 1, type=int)
        per_page = request.args.get("per_page", 10, type=int)
        branch_id = request.args.get("branch_id", type=int)
        warehouse_id = request.args.get("warehouse_id", type=int)
        date_from = request.args.get("date_from")
        date_to = request.args.get("date_to")
        search = request.args.get("search", "")
        
        # Validate pagination parameters
        if page < 1:
            return create_error_response("Page number must be positive", 400, "page")
        if per_page < 1 or per_page > 100:
            return create_error_response("Items per page must be between 1 and 100", 400, "per_page")

        try:
            # Build query with joins and explicit relationship loading
            from sqlalchemy.orm import joinedload
            query = (
                Transaction.query
                .options(
                    noload(Transaction.asset_transactions), 
                    noload(Transaction.user),
                    joinedload(Transaction.warehouse)  # Explicitly load warehouse relationship
                )
                .join(Warehouse)
            )
            # Apply filters
            if branch_id:
                query = query.filter(Warehouse.branch_id == branch_id)
            
            if warehouse_id:
                query = query.filter(Transaction.warehouse_id == warehouse_id)
            
            if date_from:
                try:
                    date_from_obj = datetime.strptime(date_from, "%Y-%m-%d").date()
                    query = query.filter(Transaction.date >= date_from_obj)
                except ValueError:
                    return create_error_response("Invalid date_from format. Use YYYY-MM-DD", 400, "date_from")
            
            if date_to:
                try:
                    date_to_obj = datetime.strptime(date_to, "%Y-%m-%d").date()
                    query = query.filter(Transaction.date <= date_to_obj)
                except ValueError:
                    return create_error_response("Invalid date_to format. Use YYYY-MM-DD", 400, "date_to")
            
            if search:
                query = query.filter(or_(
                    Transaction.description.contains(search),
                    Transaction.reference_number.contains(search)
                ))

            # Order by ID descending for consistent ordering
            query = query.order_by(Transaction.id.desc())

            paginated = query.paginate(page=page, per_page=per_page, error_out=False)
            return {
                "items": transactions_schema.dump(paginated.items),
                "total": paginated.total,
                "page": paginated.page,
                "pages": paginated.pages
            }
        except OperationalError as e:
            logging.error(f"Database operational error in transaction list: {str(e)}")
            return create_error_response("Database connection error", 503)
        except SQLAlchemyError as e:
            logging.error(f"Database error in transaction list: {str(e)}")
            return create_error_response("Database error occurred", 500)
        except Exception as e:
            logging.error(f"Unexpected error in transaction list: {str(e)}")
            return create_error_response("Internal server error", 500)
    

    @transactions_ns.doc('create_transaction', security='Bearer Auth')
    #@transactions_ns.expect(transaction_create_model)
    @transactions_ns.expect(file_upload_parser)
    @transactions_ns.response(201, 'Successfully created transaction', transaction_model)
    @transactions_ns.response(400, 'Bad Request', error_model)
    @transactions_ns.response(401, 'Unauthorized', error_model)
    @transactions_ns.response(403, 'Forbidden', error_model)
    @transactions_ns.response(404, 'Warehouse/Asset not found', error_model)
    @transactions_ns.response(409, 'Conflict - Insufficient quantity', error_model)
    @transactions_ns.response(500, 'Internal Server Error', error_model)
    @transactions_ns.response(503, 'Service Unavailable', error_model)
    @jwt_required()
    def post(self):
        """Create a new transaction with asset transactions"""
        error = check_permission("can_make_transaction")
        if error:
            return error

        try:
            print("Starting transaction creation process")
            
            # Handle file upload
            attached_file_name = None
            
            # Check if request has files
            if 'attached_file' in request.files:
                file = request.files['attached_file']
                
                if file and file.filename and file.filename.strip():
                    try:
                        print(f"Processing file: {file.filename}")
                        # Generate unique filename
                        file_extension = os.path.splitext(secure_filename(file.filename))[1]
                        unique_filename = f"{uuid.uuid4().hex}{file_extension}"
                        
                        # Create upload directory if it doesn't exist
                        upload_folder = current_app.config.get('UPLOAD_FOLDER', 'uploads')
                        os.makedirs(upload_folder, exist_ok=True)
                        
                        # Save file
                        file_path = os.path.join(upload_folder, unique_filename)
                        file.save(file_path)
                        attached_file_name = unique_filename
                        print(f"File saved as: {unique_filename}")
                        
                    except Exception as e:
                        print(f"File upload error: {str(e)}")
                        return create_error_response(f"File upload failed: {str(e)}", 400)
            
            # Handle JSON data
            try:
                # For multipart requests, JSON data might be in a form field
                if request.content_type and 'multipart/form-data' in request.content_type:
                    # Try to get JSON from form data
                    json_str = request.form.get('data')
                    if json_str:
                        json_data = json.loads(json_str)
                    else:
                        return create_error_response("No transaction data provided in multipart request", 400)
                else:
                    # Regular JSON request
                    json_data = request.get_json()
                    if not json_data:
                        return create_error_response("No JSON data provided", 400)
                
                data = transaction_create_schema.load(json_data)
                print("Transaction data validated successfully")
                
            except json.JSONDecodeError as e:
                print(f"JSON decode error: {str(e)}")
                return create_error_response(f"Invalid JSON format: {str(e)}", 400)
            except ValidationError as e:
                print(f"Schema validation error: {str(e)}")
                return create_validation_error_response(e.messages)
            except Exception as e:
                print(f"Data processing error: {str(e)}")
                return create_error_response(f"Data processing failed: {str(e)}", 400)
            
            # Get warehouse to determine branch
            warehouse = db.session.get(Warehouse, data['warehouse_id'])
            if not warehouse:
                return create_error_response("Warehouse not found", 404, "warehouse_id")
            
            # Generate custom_id
            custom_id = Transaction.generate_custom_id(warehouse.branch_id)
            
            # Create transaction
            transaction_data = {
                'custom_id': custom_id,
                'date': data['date'],
                'description': data.get('description'),
                'reference_number': data.get('reference_number'),
                'user_id': get_jwt_identity(),  # Assuming you want to track the user creating the transaction
                'warehouse_id': data['warehouse_id'],
                'transaction_type': data['transaction_type'],  # Now at transaction level
                'attached_file': attached_file_name  # Store the unique filename
            }
            
            new_transaction = Transaction(**transaction_data)
            db.session.add(new_transaction)
            db.session.flush()  # Get the transaction ID
            
            # Create asset transactions and update asset quantities
            asset_transactions = []
            transaction_type = data['transaction_type']  # Get transaction type once
            for asset_trans_data in data['asset_transactions']:
                # Get the asset
                asset = db.session.get(FixedAsset, asset_trans_data['asset_id'])
                if not asset:
                    db.session.rollback()
                    return create_error_response(f"Asset {asset_trans_data['asset_id']} not found", 404, "asset_id")
                
                # Check if OUT transaction has enough quantity
                if not transaction_type:  # OUT transaction
                    if asset.quantity < asset_trans_data['quantity']:
                        beforeRollback = asset.quantity 
                        db.session.rollback()
                        afterRollback = asset.quantity - beforeRollback
                        return create_error_response(
                            f"Insufficient quantity for asset {asset.name_en}. "
                            f"Available: {asset.quantity}, Requested: {asset_trans_data['quantity'] + afterRollback}",
                            409, "quantity"
                        )
                
                # Update asset quantity
                if transaction_type:  # IN transaction
                    asset.quantity += asset_trans_data['quantity']
                else:  # OUT transaction
                    asset.quantity -= asset_trans_data['quantity']
                
                # Create asset transaction
                asset_trans = AssetTransaction(
                    transaction_id=new_transaction.id,
                    asset_id=asset_trans_data['asset_id'],
                    quantity=asset_trans_data['quantity'],
                    amount=asset_trans_data.get('amount')
                )
                # Calculate total_value automatically via the model's __init__
                asset_transactions.append(asset_trans)
            
            db.session.add_all(asset_transactions)
            db.session.commit()
            
            return transaction_schema.dump(new_transaction), 201
            
        except IntegrityError as e:
            db.session.rollback()
            logging.error(f"Integrity error creating transaction: {str(e)}")
            return create_error_response("Data integrity constraint violation", 409)
        except OperationalError as e:
            db.session.rollback()
            logging.error(f"Database operational error creating transaction: {str(e)}")
            return create_error_response("Database connection error", 503)
        except SQLAlchemyError as e:
            db.session.rollback()
            logging.error(f"Database error creating transaction: {str(e)}")
            return create_error_response("Database error occurred", 500)
        except Exception as e:
            db.session.rollback()
            logging.error(f"Unexpected error creating transaction: {str(e)}")
            return create_error_response("Internal server error", 500)


@transactions_ns.route("/<int:transaction_id>")
class TransactionResource(Resource):
    @transactions_ns.doc('get_transaction', security='Bearer Auth')
    @transactions_ns.response(200, 'Successfully retrieved transaction', transaction_model)
    @transactions_ns.response(401, 'Unauthorized', error_model)
    @transactions_ns.response(403, 'Forbidden', error_model)
    @transactions_ns.response(404, 'Transaction not found', error_model)
    @transactions_ns.response(500, 'Internal Server Error', error_model)
    @transactions_ns.response(503, 'Service Unavailable', error_model)
    @jwt_required()
    def get(self, transaction_id):
        """Get a specific transaction"""
        error = check_permission("can_make_transaction")
        if error:
            return error

        try:
            transaction = db.session.get(Transaction, transaction_id)
            if not transaction:
                return create_error_response("Transaction not found", 404)
            return transaction_schema.dump(transaction)
        except OperationalError as e:
            logging.error(f"Database operational error getting transaction {transaction_id}: {str(e)}")
            return create_error_response("Database connection error", 503)
        except SQLAlchemyError as e:
            logging.error(f"Database error getting transaction {transaction_id}: {str(e)}")
            return create_error_response("Database error occurred", 500)
        except Exception as e:
            logging.error(f"Unexpected error getting transaction {transaction_id}: {str(e)}")
            return create_error_response("Internal server error", 500)




    @transactions_ns.doc('update_transaction', security='Bearer Auth')
    @transactions_ns.expect(transaction_input_model)
    @transactions_ns.response(200, 'Successfully updated transaction', transaction_model)
    @transactions_ns.response(400, 'Bad Request', error_model)
    @transactions_ns.response(401, 'Unauthorized', error_model)
    @transactions_ns.response(403, 'Forbidden', error_model)
    @transactions_ns.response(404, 'Transaction not found', error_model)
    @transactions_ns.response(500, 'Internal Server Error', error_model)
    @transactions_ns.response(503, 'Service Unavailable', error_model)
    @jwt_required()
    def put(self, transaction_id):
        """Update a transaction"""
        error = check_permission("can_make_transaction")
        if error:
            return error

        # Validate request body
        json_data = request.get_json()
        if not json_data:
            return create_error_response("Request body is required", 400)

        try:
            transaction = db.session.get(Transaction, transaction_id)
            if not transaction:
                return create_error_response("Transaction not found", 404)

            try:
                data = transaction_schema.load(json_data, partial=True)
            except ValidationError as err:
                return create_validation_error_response(err.messages)
            
            # Don't allow updating custom_id
            if 'custom_id' in data:
                del data['custom_id']
            
            for key, value in data.items():
                setattr(transaction, key, value)
            
            db.session.commit()
            return transaction_schema.dump(transaction)
        except IntegrityError as e:
            db.session.rollback()
            logging.error(f"Integrity error updating transaction {transaction_id}: {str(e)}")
            return create_error_response("Data integrity constraint violation", 409)
        except OperationalError as e:
            db.session.rollback()
            logging.error(f"Database operational error updating transaction {transaction_id}: {str(e)}")
            return create_error_response("Database connection error", 503)
        except SQLAlchemyError as e:
            db.session.rollback()
            logging.error(f"Database error updating transaction {transaction_id}: {str(e)}")
            return create_error_response("Database error occurred", 500)
        except Exception as e:
            db.session.rollback()
            logging.error(f"Unexpected error updating transaction {transaction_id}: {str(e)}")
            return create_error_response("Internal server error", 500)

    @transactions_ns.doc('delete_transaction', security='Bearer Auth')
    @transactions_ns.response(200, 'Successfully deleted transaction', success_model)
    @transactions_ns.response(401, 'Unauthorized', error_model)
    @transactions_ns.response(403, 'Forbidden', error_model)
    @transactions_ns.response(404, 'Transaction not found', error_model)
    @transactions_ns.response(500, 'Internal Server Error', error_model)
    @transactions_ns.response(503, 'Service Unavailable', error_model)
    @jwt_required()
    def delete(self, transaction_id):
        """Delete a transaction"""
        error = check_permission("can_make_transaction")
        if error:
            return error

        try:
            transaction = db.session.get(Transaction, transaction_id)
            if not transaction:
                return create_error_response("Transaction not found", 404)

            # Before deleting transaction, reverse all asset quantity changes
            transaction_type = transaction.transaction_type  # Get transaction type once
            for asset_trans in transaction.asset_transactions:
                asset = db.session.get(FixedAsset, asset_trans.asset_id)
                if asset:
                    if transaction_type:  # Was IN transaction
                        asset.quantity -= asset_trans.quantity  # Remove the added quantity
                    else:  # Was OUT transaction
                        asset.quantity += asset_trans.quantity  # Add back the removed quantity
            
            db.session.delete(transaction)  # Cascade will delete asset_transactions
            db.session.commit()
            return {"message": f"Transaction {transaction_id} deleted successfully"}
        except IntegrityError as e:
            db.session.rollback()
            logging.error(f"Integrity error deleting transaction {transaction_id}: {str(e)}")
            return create_error_response("Cannot delete transaction: it may be referenced by other records", 409)
        except OperationalError as e:
            db.session.rollback()
            logging.error(f"Database operational error deleting transaction {transaction_id}: {str(e)}")
            return create_error_response("Database connection error", 503)
        except SQLAlchemyError as e:
            db.session.rollback()
            logging.error(f"Database error deleting transaction {transaction_id}: {str(e)}")
            return create_error_response("Database error occurred", 500)
        except Exception as e:
            db.session.rollback()
            logging.error(f"Unexpected error deleting transaction {transaction_id}: {str(e)}")
            return create_error_response("Internal server error", 500)


@transactions_ns.route("/<int:transaction_id>/assets")
class TransactionAssetsList(Resource):
    @transactions_ns.doc('get_transaction_assets', security='Bearer Auth')
    @transactions_ns.response(200, 'Successfully retrieved transaction assets', pagination_model)
    @transactions_ns.response(400, 'Bad Request', error_model)
    @transactions_ns.response(401, 'Unauthorized', error_model)
    @transactions_ns.response(403, 'Forbidden', error_model)
    @transactions_ns.response(404, 'Transaction not found', error_model)
    @transactions_ns.response(500, 'Internal Server Error', error_model)
    @transactions_ns.response(503, 'Service Unavailable', error_model)
    @transactions_ns.param('page', 'Page number', type=int, default=1)
    @transactions_ns.param('per_page', 'Items per page', type=int, default=10)
    @jwt_required()
    def get(self, transaction_id):
        """Get all asset transactions for a specific transaction"""
        error = check_permission("can_make_transaction")
        if error:
            return error

        page = request.args.get("page", 1, type=int)
        per_page = request.args.get("per_page", 10, type=int)
        
        # Validate pagination parameters
        if page < 1:
            return create_error_response("Page number must be positive", 400, "page")
        if per_page < 1 or per_page > 100:
            return create_error_response("Items per page must be between 1 and 100", 400, "per_page")

        try:
            # Check if transaction exists
            transaction = db.session.get(Transaction, transaction_id)
            if not transaction:
                return create_error_response("Transaction not found", 404)

            query = AssetTransaction.query.filter_by(transaction_id=transaction_id)
            
            # Order by ID descending for consistent ordering
            query = query.order_by(AssetTransaction.id.desc())

            paginated = query.paginate(page=page, per_page=per_page, error_out=False)
            return {
                "items": asset_transactions_schema.dump(paginated.items),
                "total": paginated.total,
                "page": paginated.page,
                "pages": paginated.pages
            }
        except OperationalError as e:
            logging.error(f"Database operational error getting transaction assets {transaction_id}: {str(e)}")
            return create_error_response("Database connection error", 503)
        except SQLAlchemyError as e:
            logging.error(f"Database error getting transaction assets {transaction_id}: {str(e)}")
            return create_error_response("Database error occurred", 500)
        except Exception as e:
            logging.error(f"Unexpected error getting transaction assets {transaction_id}: {str(e)}")
            return create_error_response("Internal server error", 500)

    @transactions_ns.doc('add_asset_to_transaction', security='Bearer Auth')
    @transactions_ns.expect(asset_transaction_input_model)
    @transactions_ns.response(201, 'Successfully added asset to transaction', asset_transaction_model)
    @transactions_ns.response(400, 'Bad Request', error_model)
    @transactions_ns.response(401, 'Unauthorized', error_model)
    @transactions_ns.response(403, 'Forbidden', error_model)
    @transactions_ns.response(404, 'Transaction/Asset not found', error_model)
    @transactions_ns.response(409, 'Conflict - Insufficient quantity', error_model)
    @transactions_ns.response(500, 'Internal Server Error', error_model)
    @transactions_ns.response(503, 'Service Unavailable', error_model)
    @jwt_required()
    def post(self, transaction_id):
        """Add an asset transaction to an existing transaction"""
        error = check_permission("can_make_transaction")
        if error:
            return error

        # Validate request body
        json_data = request.get_json()
        if not json_data:
            return create_error_response("Request body is required", 400)

        try:
            # Check if transaction exists
            transaction = db.session.get(Transaction, transaction_id)
            if not transaction:
                return create_error_response("Transaction not found", 404)

            try:
                data = asset_transaction_create_schema.load(json_data)
            except ValidationError as err:
                return create_validation_error_response(err.messages)
            
            # Get the asset and check availability for OUT transactions
            asset = db.session.get(FixedAsset, data['asset_id'])
            if not asset:
                return create_error_response("Asset not found", 404, "asset_id")
            
            # Use the parent transaction's transaction_type
            transaction_type = transaction.transaction_type
            
            # Check if OUT transaction has enough quantity
            if not transaction_type:  # OUT transaction
                if asset.quantity < data['quantity']:
                    return create_error_response(
                        f"Insufficient quantity for asset {asset.name_en}. "
                        f"Available: {asset.quantity}, Requested: {data['quantity']}",
                        409, "quantity"
                    )
            
            # Update asset quantity
            if transaction_type:  # IN transaction
                asset.quantity += data['quantity']
            else:  # OUT transaction
                asset.quantity -= data['quantity']
            
            asset_transaction = AssetTransaction(
                transaction_id=transaction_id,
                asset_id=data['asset_id'],
                quantity=data['quantity'],
                amount=data.get('amount')
            )
            
            db.session.add(asset_transaction)
            db.session.commit()
            
            return asset_transaction_schema.dump(asset_transaction), 201
            
        except IntegrityError as e:
            db.session.rollback()
            logging.error(f"Integrity error adding asset to transaction {transaction_id}: {str(e)}")
            return create_error_response("Data integrity constraint violation", 409)
        except OperationalError as e:
            db.session.rollback()
            logging.error(f"Database operational error adding asset to transaction {transaction_id}: {str(e)}")
            return create_error_response("Database connection error", 503)
        except SQLAlchemyError as e:
            db.session.rollback()
            logging.error(f"Database error adding asset to transaction {transaction_id}: {str(e)}")
            return create_error_response("Database error occurred", 500)
        except Exception as e:
            db.session.rollback()
            logging.error(f"Unexpected error adding asset to transaction {transaction_id}: {str(e)}")
            return create_error_response("Internal server error", 500)


@asset_transactions_ns.route("/<int:asset_transaction_id>")
class AssetTransactionResource(Resource):
    @asset_transactions_ns.doc('get_asset_transaction', security='Bearer Auth')
    @asset_transactions_ns.response(200, 'Successfully retrieved asset transaction', asset_transaction_model)
    @asset_transactions_ns.response(401, 'Unauthorized', error_model)
    @asset_transactions_ns.response(403, 'Forbidden', error_model)
    @asset_transactions_ns.response(404, 'Asset transaction not found', error_model)
    @asset_transactions_ns.response(500, 'Internal Server Error', error_model)
    @asset_transactions_ns.response(503, 'Service Unavailable', error_model)
    @jwt_required()
    def get(self, asset_transaction_id):
        """Get a specific asset transaction"""
        error = check_permission("can_make_transaction")
        if error:
            return error

        try:
            asset_transaction = db.session.get(AssetTransaction, asset_transaction_id)
            if not asset_transaction:
                return create_error_response("Asset transaction not found", 404)
            
            return asset_transaction_schema.dump(asset_transaction)
        except OperationalError as e:
            logging.error(f"Database operational error getting asset transaction {asset_transaction_id}: {str(e)}")
            return create_error_response("Database connection error", 503)
        except SQLAlchemyError as e:
            logging.error(f"Database error getting asset transaction {asset_transaction_id}: {str(e)}")
            return create_error_response("Database error occurred", 500)
        except Exception as e:
            logging.error(f"Unexpected error getting asset transaction {asset_transaction_id}: {str(e)}")
            return create_error_response("Internal server error", 500)

    @asset_transactions_ns.doc('update_asset_transaction', security='Bearer Auth')
    @asset_transactions_ns.expect(asset_transaction_input_model)
    @asset_transactions_ns.response(200, 'Successfully updated asset transaction', asset_transaction_model)
    @asset_transactions_ns.response(400, 'Bad Request', error_model)
    @asset_transactions_ns.response(401, 'Unauthorized', error_model)
    @asset_transactions_ns.response(403, 'Forbidden', error_model)
    @asset_transactions_ns.response(404, 'Asset transaction/Asset not found', error_model)
    @asset_transactions_ns.response(409, 'Conflict - Insufficient quantity', error_model)
    @asset_transactions_ns.response(500, 'Internal Server Error', error_model)
    @asset_transactions_ns.response(503, 'Service Unavailable', error_model)
    @jwt_required()
    def put(self, asset_transaction_id):
        """Update an asset transaction"""
        error = check_permission("can_make_transaction")
        if error:
            return error

        # Validate request body
        json_data = request.get_json()
        if not json_data:
            return create_error_response("Request body is required", 400)

        try:
            asset_transaction = db.session.get(AssetTransaction, asset_transaction_id)
            if not asset_transaction:
                return create_error_response("Asset transaction not found", 404)

            try:
                data = asset_transaction_create_schema.load(json_data, partial=True)
            except ValidationError as err:
                return create_validation_error_response(err.messages)
            
            # Get current values before update
            old_quantity = asset_transaction.quantity
            old_asset_id = asset_transaction.asset_id
            
            # Get the parent transaction's transaction_type (since it's no longer in asset_transaction)
            parent_transaction = asset_transaction.transaction
            transaction_type = parent_transaction.transaction_type
            
            # Get new values
            new_quantity = data.get('quantity', old_quantity)
            new_asset_id = data.get('asset_id', old_asset_id)
            
            # If asset_id changed, we need to handle both assets
            if new_asset_id != old_asset_id:
                # Reverse the effect on the old asset
                old_asset = db.session.get(FixedAsset, old_asset_id)
                if old_asset:
                    if transaction_type:  # Was IN transaction
                        old_asset.quantity -= old_quantity  # Remove the added quantity
                    else:  # Was OUT transaction
                        old_asset.quantity += old_quantity  # Add back the removed quantity
                
                # Get the new asset
                new_asset = db.session.get(FixedAsset, new_asset_id)
                if not new_asset:
                    return create_error_response("New asset not found", 404, "asset_id")
                
                # Check availability for new asset if OUT transaction
                if not transaction_type:  # OUT transaction
                    if new_asset.quantity < new_quantity:
                        return create_error_response(
                            f"Insufficient quantity for asset {new_asset.name_en}. "
                            f"Available: {new_asset.quantity}, Requested: {new_quantity}",
                            409, "quantity"
                        )
                
                # Apply effect to new asset
                if transaction_type:  # IN transaction
                    new_asset.quantity += new_quantity
                else:  # OUT transaction
                    new_asset.quantity -= new_quantity
            
            else:
                # Same asset, but quantity might have changed
                asset = db.session.get(FixedAsset, old_asset_id)
                if asset:
                    # Reverse the old effect
                    if transaction_type:  # Was IN transaction
                        asset.quantity -= old_quantity
                    else:  # Was OUT transaction
                        asset.quantity += old_quantity
                    
                    # Check availability for OUT transaction
                    if not transaction_type:  # OUT transaction
                        if asset.quantity < new_quantity:
                            # Restore the old effect before returning error
                            if transaction_type:
                                asset.quantity += old_quantity
                            else:
                                asset.quantity -= old_quantity
                            return create_error_response(
                                f"Insufficient quantity for asset {asset.name_en}. "
                                f"Available: {asset.quantity}, Requested: {new_quantity}",
                                409, "quantity"
                            )
                    
                    # Apply the new effect
                    if transaction_type:  # IN transaction
                        asset.quantity += new_quantity
                    else:  # OUT transaction
                        asset.quantity -= new_quantity
            
            # Update asset transaction fields
            for key, value in data.items():
                setattr(asset_transaction, key, value)
            
            # Recalculate total_value if quantity or amount changed
            asset_transaction.calculate_total_value()
            
            db.session.commit()
            return asset_transaction_schema.dump(asset_transaction)
            
        except IntegrityError as e:
            db.session.rollback()
            logging.error(f"Integrity error updating asset transaction {asset_transaction_id}: {str(e)}")
            return create_error_response("Data integrity constraint violation", 409)
        except OperationalError as e:
            db.session.rollback()
            logging.error(f"Database operational error updating asset transaction {asset_transaction_id}: {str(e)}")
            return create_error_response("Database connection error", 503)
        except SQLAlchemyError as e:
            db.session.rollback()
            logging.error(f"Database error updating asset transaction {asset_transaction_id}: {str(e)}")
            return create_error_response("Database error occurred", 500)
        except Exception as e:
            db.session.rollback()
            logging.error(f"Unexpected error updating asset transaction {asset_transaction_id}: {str(e)}")
            return create_error_response("Internal server error", 500)

    @asset_transactions_ns.doc('delete_asset_transaction', security='Bearer Auth')
    @asset_transactions_ns.response(200, 'Successfully deleted asset transaction', success_model)
    @asset_transactions_ns.response(401, 'Unauthorized', error_model)
    @asset_transactions_ns.response(403, 'Forbidden', error_model)
    @asset_transactions_ns.response(404, 'Asset transaction not found', error_model)
    @asset_transactions_ns.response(500, 'Internal Server Error', error_model)
    @asset_transactions_ns.response(503, 'Service Unavailable', error_model)
    @jwt_required()
    def delete(self, asset_transaction_id):
        """Delete an asset transaction"""
        error = check_permission("can_make_transaction")
        if error:
            return error

        try:
            asset_transaction = db.session.get(AssetTransaction, asset_transaction_id)
            if not asset_transaction:
                return create_error_response("Asset transaction not found", 404)

            # Before deleting, reverse the quantity effect on the asset
            # Get the parent transaction's transaction_type
            parent_transaction = asset_transaction.transaction
            transaction_type = parent_transaction.transaction_type
            
            asset = db.session.get(FixedAsset, asset_transaction.asset_id)
            if asset:
                if transaction_type:  # Was IN transaction
                    asset.quantity -= asset_transaction.quantity  # Remove the added quantity
                else:  # Was OUT transaction
                    asset.quantity += asset_transaction.quantity  # Add back the removed quantity
            
            db.session.delete(asset_transaction)
            db.session.commit()
            return {"message": f"Asset transaction {asset_transaction_id} deleted successfully"}
        except IntegrityError as e:
            db.session.rollback()
            logging.error(f"Integrity error deleting asset transaction {asset_transaction_id}: {str(e)}")
            return create_error_response("Cannot delete asset transaction: it may be referenced by other records", 409)
        except OperationalError as e:
            db.session.rollback()
            logging.error(f"Database operational error deleting asset transaction {asset_transaction_id}: {str(e)}")
            return create_error_response("Database connection error", 503)
        except SQLAlchemyError as e:
            db.session.rollback()
            logging.error(f"Database error deleting asset transaction {asset_transaction_id}: {str(e)}")
            return create_error_response("Database error occurred", 500)
        except Exception as e:
            db.session.rollback()
            logging.error(f"Unexpected error deleting asset transaction {asset_transaction_id}: {str(e)}")
            return create_error_response("Internal server error", 500)


@transactions_ns.route("/summary")
class TransactionSummary(Resource):
    @transactions_ns.doc('get_transaction_summary', security='Bearer Auth')
    @transactions_ns.response(200, 'Successfully retrieved transaction summary')
    @transactions_ns.response(400, 'Bad Request', error_model)
    @transactions_ns.response(401, 'Unauthorized', error_model)
    @transactions_ns.response(403, 'Forbidden', error_model)
    @transactions_ns.response(500, 'Internal Server Error', error_model)
    @transactions_ns.response(503, 'Service Unavailable', error_model)
    @transactions_ns.param('branch_id', 'Filter by branch ID', type=int)
    @transactions_ns.param('warehouse_id', 'Filter by warehouse ID', type=int)
    @transactions_ns.param('date_from', 'Summary from date (YYYY-MM-DD)', type=str)
    @transactions_ns.param('date_to', 'Summary to date (YYYY-MM-DD)', type=str)
    @jwt_required()
    def get(self):
        """Get transaction summary statistics"""
        error = check_permission("can_make_transaction")
        if error:
            return error

        branch_id = request.args.get("branch_id", type=int)
        warehouse_id = request.args.get("warehouse_id", type=int)
        date_from = request.args.get("date_from")
        date_to = request.args.get("date_to")

        try:
            # Build base query
            transaction_query = Transaction.query.join(Warehouse)
            asset_transaction_query = AssetTransaction.query.join(Transaction).join(Warehouse)

            # Apply filters
            if branch_id:
                transaction_query = transaction_query.filter(Warehouse.branch_id == branch_id)
                asset_transaction_query = asset_transaction_query.filter(Warehouse.branch_id == branch_id)
            
            if warehouse_id:
                transaction_query = transaction_query.filter(Transaction.warehouse_id == warehouse_id)
                asset_transaction_query = asset_transaction_query.filter(Transaction.warehouse_id == warehouse_id)
            
            if date_from:
                try:
                    date_from_obj = datetime.strptime(date_from, "%Y-%m-%d").date()
                    transaction_query = transaction_query.filter(Transaction.date >= date_from_obj)
                    asset_transaction_query = asset_transaction_query.filter(Transaction.date >= date_from_obj)
                except ValueError:
                    return create_error_response("Invalid date_from format. Use YYYY-MM-DD", 400, "date_from")
            
            if date_to:
                try:
                    date_to_obj = datetime.strptime(date_to, "%Y-%m-%d").date()
                    transaction_query = transaction_query.filter(Transaction.date <= date_to_obj)
                    asset_transaction_query = asset_transaction_query.filter(Transaction.date <= date_to_obj)
                except ValueError:
                    return create_error_response("Invalid date_to format. Use YYYY-MM-DD", 400, "date_to")

            # Calculate statistics
            total_transactions = transaction_query.count()
            total_in_transactions = transaction_query.filter(Transaction.transaction_type == True).count()
            total_out_transactions = transaction_query.filter(Transaction.transaction_type == False).count()
            
            # Calculate total values
            from sqlalchemy import func
            total_in_value = asset_transaction_query.filter(
                Transaction.transaction_type == True,
                AssetTransaction.total_value.isnot(None)
            ).with_entities(func.sum(AssetTransaction.total_value)).scalar() or 0
            
            total_out_value = asset_transaction_query.filter(
                Transaction.transaction_type == False,
                AssetTransaction.total_value.isnot(None)
            ).with_entities(func.sum(AssetTransaction.total_value)).scalar() or 0

            return {
                "total_transactions": total_transactions,
                "total_in_transactions": total_in_transactions,
                "total_out_transactions": total_out_transactions,
                "total_in_value": float(total_in_value),
                "total_out_value": float(total_out_value),
                "net_value": float(total_in_value - total_out_value)
            }
        except OperationalError as e:
            logging.error(f"Database operational error in transaction summary: {str(e)}")
            return create_error_response("Database connection error", 503)
        except SQLAlchemyError as e:
            logging.error(f"Database error in transaction summary: {str(e)}")
            return create_error_response("Database error occurred", 500)
        except Exception as e:
            logging.error(f"Unexpected error in transaction summary: {str(e)}")
            return create_error_response("Internal server error", 500)
    

# New Resource for downloading transaction file
@transactions_ns.route('/<int:transaction_id>/download')
class TransactionDownloadResource(Resource):
    @transactions_ns.doc('download_transaction_file', security='Bearer Auth')
    @transactions_ns.response(200, 'Successfully downloaded file')
    @transactions_ns.response(401, 'Unauthorized', error_model)
    @transactions_ns.response(403, 'Forbidden', error_model)
    @transactions_ns.response(404, 'Transaction/File not found', error_model)
    @transactions_ns.response(500, 'Internal Server Error', error_model)
    @transactions_ns.response(503, 'Service Unavailable', error_model)
    @transactions_ns.param('token', 'JWT token for authentication (alternative to Authorization header)', type=str)
    def get(self, transaction_id):
        """Download the attached file for a transaction"""
        # Authenticate user using custom function that handles both header and query token
        user_id = authenticate_for_download()
        
        if not user_id:
            return create_error_response("Missing or invalid authentication. Provide Authorization header or token parameter.", 401)
        
        # Check user permissions
        from ..models import User
        user = User.query.get(user_id)
        if not user:
            return create_error_response("User not found", 404)
            
        if not getattr(user, "can_make_transaction", False):
            return create_error_response("Permission 'can_make_transaction' denied", 403)

        try:
            transaction = db.session.get(Transaction, transaction_id)
            if not transaction:
                return create_error_response("Transaction not found", 404)

            if not transaction.attached_file:
                return create_error_response("No file attached to this transaction.", 404)

            file_path = transaction.attached_file
            # If file path is not absolute, assume uploads folder
            if not os.path.isabs(file_path):
                file_path = os.path.join(current_app.root_path, '..', 'uploads', file_path)

            if not os.path.exists(file_path):
                return create_error_response("File not found.", 404)

            return send_file(file_path, as_attachment=True)
        except OperationalError as e:
            logging.error(f"Database operational error downloading file for transaction {transaction_id}: {str(e)}")
            return create_error_response("Database connection error", 503)
        except SQLAlchemyError as e:
            logging.error(f"Database error downloading file for transaction {transaction_id}: {str(e)}")
            return create_error_response("Database error occurred", 500)
        except Exception as e:
            logging.error(f"Unexpected error downloading file for transaction {transaction_id}: {str(e)}")
            return create_error_response("Internal server error", 500)


@transactions_ns.route("/generate-report")
class GenerateReport(Resource):
    @transactions_ns.doc('generate_report', security='Bearer Auth')
    @transactions_ns.param('date', 'Filter by exact date (YYYY-MM-DD) - REQUIRED', type=str, required=True)
    @transactions_ns.param('category', 'Filter by category name', type=str)
    @transactions_ns.param('subcategory', 'Filter by subcategory name', type=str)
    @transactions_ns.param('branch_id', 'Filter by branch ID', type=int)
    @transactions_ns.param('warehouse_id', 'Filter by warehouse ID', type=int)
    @jwt_required()
    def get(self):
        """Generate comprehensive transaction report with asset-level analysis
        
        Returns asset-wise summary including:
        - Asset name and details
        - Total quantity IN and OUT
        - Total amount IN and OUT  
        - Total cost (quantity × amount) IN and OUT
        - Overall totals
        
        REQUIRED: date parameter must be provided
        Optional filters can be combined:
        - Category: ?category=Electronics
        - Subcategory: ?subcategory=Laptops (can be used with or without category)
        - Branch/Warehouse: ?branch_id=1 or ?warehouse_id=1
        """
        error = check_permission("can_make_report")
        if error:
            return error

        # Get filter parameters - date is REQUIRED
        exact_date = request.args.get("date")
        if not exact_date:
            return create_error_response("Date parameter is required. Use format: YYYY-MM-DD", 400, "date")

        category_name = request.args.get("category")
        subcategory_name = request.args.get("subcategory")
        branch_id = request.args.get("branch_id", type=int)
        warehouse_id = request.args.get("warehouse_id", type=int)

        try:
            # Step 1: Parse and validate the required date
            try:
                date_obj = datetime.strptime(exact_date, "%Y-%m-%d").date()
            except ValueError:
                return create_error_response("Invalid date format. Use YYYY-MM-DD", 400, "date")

            # Step 2: Start with transactions filtered by the specific date
            print(f"Filtering transactions for date: {date_obj}")
            transaction_query = db.session.query(Transaction).filter(Transaction.date == date_obj)
            
            # Step 3: Apply warehouse filter first (most specific)
            if warehouse_id:
                print(f"Filtering by warehouse_id: {warehouse_id}")
                transaction_query = transaction_query.filter(Transaction.warehouse_id == warehouse_id)
            
            # Step 4: If no warehouse filter, but branch filter exists, apply it
            elif branch_id:
                print(f"Filtering by branch_id: {branch_id}")
                transaction_query = transaction_query.join(Warehouse).filter(Warehouse.branch_id == branch_id)
            
            # Step 5: Get the filtered transaction IDs (this limits our scope early)
            filtered_transaction_ids = [t.id for t in transaction_query.all()]
            
            if not filtered_transaction_ids:
                print("No transactions found for the given filters")
                return {
                    'report_metadata': {
                        'generated_at': datetime.now().isoformat(),
                        'filters_applied': {
                            'date': exact_date,
                            'category': category_name,
                            'subcategory': subcategory_name,
                            'branch_id': branch_id,
                            'warehouse_id': warehouse_id
                        },
                        'total_assets': 0
                    },
                    'asset_reports': [],
                    'summary_totals': {
                        'total_quantity_in': 0,
                        'total_quantity_out': 0,
                        'total_amount_in': 0,
                        'total_amount_out': 0,
                        'total_cost_in': 0,
                        'total_cost_out': 0,
                        'net_quantity': 0,
                        'net_amount': 0,
                        'net_cost': 0
                    }
                }, 200

            print(f"Found {len(filtered_transaction_ids)} transactions to analyze")

            # Step 6: Build optimized query using the filtered transaction IDs
            from sqlalchemy import func, case
            query = db.session.query(
                FixedAsset.id.label('asset_id'),
                FixedAsset.name_ar.label('asset_name_ar'),
                FixedAsset.name_en.label('asset_name_en'),
                FixedAsset.product_code.label('product_code'),
                Category.category.label('category'),
                Category.subcategory.label('subcategory'),
                # Sum quantities for IN transactions (transaction_type = True)
                func.coalesce(
                    func.sum(
                        case(
                            (Transaction.transaction_type == True, AssetTransaction.quantity),
                            else_=0
                        )
                    ), 0
                ).label('total_quantity_in'),
                # Sum quantities for OUT transactions (transaction_type = False)
                func.coalesce(
                    func.sum(
                        case(
                            (Transaction.transaction_type == False, AssetTransaction.quantity),
                            else_=0
                        )
                    ), 0
                ).label('total_quantity_out'),
                # Sum amounts for IN transactions
                func.coalesce(
                    func.sum(
                        case(
                            (Transaction.transaction_type == True, AssetTransaction.amount),
                            else_=0
                        )
                    ), 0
                ).label('total_amount_in'),
                # Sum amounts for OUT transactions
                func.coalesce(
                    func.sum(
                        case(
                            (Transaction.transaction_type == False, AssetTransaction.amount),
                            else_=0
                        )
                    ), 0
                ).label('total_amount_out'),
                # Sum total costs for IN transactions (quantity × amount)
                func.coalesce(
                    func.sum(
                        case(
                            (Transaction.transaction_type == True, AssetTransaction.quantity * AssetTransaction.amount),
                            else_=0
                        )
                    ), 0
                ).label('total_cost_in'),
                # Sum total costs for OUT transactions
                func.coalesce(
                    func.sum(
                        case(
                            (Transaction.transaction_type == False, AssetTransaction.quantity * AssetTransaction.amount),
                            else_=0
                        )
                    ), 0
                ).label('total_cost_out')
            ).select_from(
                AssetTransaction
            ).join(
                Transaction, AssetTransaction.transaction_id == Transaction.id
            ).join(
                FixedAsset, AssetTransaction.asset_id == FixedAsset.id
            ).join(
                Category, FixedAsset.category_id == Category.id
            ).filter(
                # Only include asset transactions from our filtered transactions
                AssetTransaction.transaction_id.in_(filtered_transaction_ids)
            )

            # Step 7: Apply category and subcategory filters if provided
            if category_name:
                print(f"Filtering by category: {category_name}")
                query = query.filter(Category.category == category_name)
            
            if subcategory_name:
                print(f"Filtering by subcategory: {subcategory_name}")
                query = query.filter(Category.subcategory == subcategory_name)

            # Step 8: Group by asset to get asset-level aggregations
            query = query.group_by(
                FixedAsset.id,
                FixedAsset.name_ar,
                FixedAsset.name_en,
                FixedAsset.product_code,
                Category.category,
                Category.subcategory
            )

            # Step 9: Execute the optimized query
            print("Executing final aggregation query...")
            results = query.all()
            print(f"Found {len(results)} assets with transactions")

            # Step 10: Process results
            asset_reports = []
            total_summary = {
                'total_quantity_in': 0,
                'total_quantity_out': 0,
                'total_amount_in': 0,
                'total_amount_out': 0,
                'total_cost_in': 0,
                'total_cost_out': 0
            }

            for result in results:
                asset_data = {
                    'asset_id': result.asset_id,
                    'asset_name_ar': result.asset_name_ar,
                    'asset_name_en': result.asset_name_en,
                    'product_code': result.product_code,
                    'category': result.category,
                    'subcategory': result.subcategory,
                    'total_quantity_in': int(result.total_quantity_in),
                    'total_quantity_out': int(result.total_quantity_out),
                    'total_amount_in': float(result.total_amount_in),
                    'total_amount_out': float(result.total_amount_out),
                    'total_cost_in': float(result.total_cost_in),
                    'total_cost_out': float(result.total_cost_out),
                    'net_quantity': int(result.total_quantity_in - result.total_quantity_out),
                    'net_amount': float(result.total_amount_in - result.total_amount_out),
                    'net_cost': float(result.total_cost_in - result.total_cost_out)
                }
                
                asset_reports.append(asset_data)
                
                # Add to totals
                total_summary['total_quantity_in'] += asset_data['total_quantity_in']
                total_summary['total_quantity_out'] += asset_data['total_quantity_out']
                total_summary['total_amount_in'] += asset_data['total_amount_in']
                total_summary['total_amount_out'] += asset_data['total_amount_out']
                total_summary['total_cost_in'] += asset_data['total_cost_in']
                total_summary['total_cost_out'] += asset_data['total_cost_out']

            # Add net totals
            total_summary['net_quantity'] = total_summary['total_quantity_in'] - total_summary['total_quantity_out']
            total_summary['net_amount'] = total_summary['total_amount_in'] - total_summary['total_amount_out']
            total_summary['net_cost'] = total_summary['total_cost_in'] - total_summary['total_cost_out']

            # Build response
            response = {
                'report_metadata': {
                    'generated_at': datetime.now().isoformat(),
                    'filters_applied': {
                        'date': exact_date,
                        'category': category_name,
                        'subcategory': subcategory_name,
                        'branch_id': branch_id,
                        'warehouse_id': warehouse_id
                    },
                    'total_assets': len(asset_reports),
                    'total_transactions_analyzed': len(filtered_transaction_ids)
                },
                'asset_reports': asset_reports,
                'summary_totals': total_summary
            }

            print(f"Report generated successfully with {len(asset_reports)} assets")
            return response, 200

        except Exception as e:
            print(f"Report generation error: {str(e)}")
            return create_error_response(f"Report generation failed: {str(e)}", 500)


@transactions_ns.route("/asset-average/<int:asset_id>")
class AssetAverage(Resource):
    @transactions_ns.response(200, 'Successfully retrieved asset average')
    @transactions_ns.response(401, 'Unauthorized', error_model)
    @transactions_ns.response(403, 'Forbidden', error_model)
    @transactions_ns.response(500, 'Internal Server Error', error_model)
    @transactions_ns.response(503, 'Service Unavailable', error_model)
    @jwt_required()
    def get(self, asset_id):
        """Get average cost for asset from IN transactions"""
        # Simple permission check
        error = check_permission("can_make_report")
        if error:
            return error
        
        try:
            # Use SQLAlchemy ORM instead of raw SQL
            from sqlalchemy import func
            
            # Query using ORM - much cleaner than raw SQL
            result = db.session.query(
                func.avg(AssetTransaction.amount)
            ).join(
                Transaction, AssetTransaction.transaction_id == Transaction.id
            ).filter(
                AssetTransaction.asset_id == asset_id,
                Transaction.transaction_type == True,  # IN transactions
                AssetTransaction.amount > 0
            ).scalar()
            
            if result:
                return {"asset_id": asset_id, "average": round(float(result), 2)}
            else:
                return {"asset_id": asset_id, "average": 0.0}
        except OperationalError as e:
            logging.error(f"Database operational error getting asset average for {asset_id}: {str(e)}")
            return create_error_response("Database connection error", 503)
        except SQLAlchemyError as e:
            logging.error(f"Database error getting asset average for {asset_id}: {str(e)}")
            return create_error_response("Database error occurred", 500)
        except Exception as e:
            logging.error(f"Unexpected error getting asset average for {asset_id}: {str(e)}")
            return create_error_response("Internal server error", 500)



@transactions_ns.route("/generate-excel-report")
class GenerateExcelReport(Resource):
    @transactions_ns.doc('generate_excel_report', security='Bearer Auth')
    @transactions_ns.param('date', 'Filter by exact date (YYYY-MM-DD) - REQUIRED', type=str, required=True)
    @transactions_ns.param('category', 'Filter by category name', type=str)
    @transactions_ns.param('subcategory', 'Filter by subcategory name', type=str)
    @transactions_ns.param('branch_id', 'Filter by branch ID', type=int)
    @transactions_ns.param('warehouse_id', 'Filter by warehouse ID', type=int)
    @transactions_ns.response(200, 'Successfully generated Excel report')
    @transactions_ns.response(400, 'Bad Request', error_model)
    @transactions_ns.response(401, 'Unauthorized', error_model)
    @transactions_ns.response(403, 'Forbidden', error_model)
    @transactions_ns.response(500, 'Internal Server Error', error_model)
    @jwt_required()
    def get(self):
        """Generate comprehensive transaction report as Excel file
        
        Returns Excel file with:
        - Filter information at the top
        - Asset-wise summary including:
          * Asset name and details
          * Total quantity IN and OUT
          * Total amount IN and OUT  
          * Total cost (quantity × amount) IN and OUT
          * Overall totals
        
        REQUIRED: date parameter must be provided
        Optional filters can be combined:
        - Category: ?category=Electronics
        - Subcategory: ?subcategory=Laptops (can be used with or without category)
        - Branch/Warehouse: ?branch_id=1 or ?warehouse_id=1
        """
        error = check_permission("can_make_report")
        if error:
            return error

        # Get filter parameters - date is REQUIRED
        exact_date = request.args.get("date")
        if not exact_date:
            return create_error_response("Date parameter is required. Use format: YYYY-MM-DD", 400, "date")

        category_name = request.args.get("category")
        subcategory_name = request.args.get("subcategory")
        branch_id = request.args.get("branch_id", type=int)
        warehouse_id = request.args.get("warehouse_id", type=int)

        try:
            # Step 1: Parse and validate the required date
            try:
                date_obj = datetime.strptime(exact_date, "%Y-%m-%d").date()
            except ValueError:
                return create_error_response("Invalid date format. Use YYYY-MM-DD", 400, "date")

            # Step 2: Start with transactions filtered by the specific date
            print(f"Filtering transactions for date: {date_obj}")
            transaction_query = db.session.query(Transaction).filter(Transaction.date == date_obj)
            
            # Step 3: Apply warehouse filter first (most specific)
            if warehouse_id:
                print(f"Filtering by warehouse_id: {warehouse_id}")
                transaction_query = transaction_query.filter(Transaction.warehouse_id == warehouse_id)
            
            # Step 4: If no warehouse filter, but branch filter exists, apply it
            elif branch_id:
                print(f"Filtering by branch_id: {branch_id}")
                transaction_query = transaction_query.join(Warehouse).filter(Warehouse.branch_id == branch_id)
            
            # Step 5: Get the filtered transaction IDs (this limits our scope early)
            filtered_transaction_ids = [t.id for t in transaction_query.all()]
            
            if not filtered_transaction_ids:
                print("No transactions found for the given filters")
                # Create empty Excel file with filter info only
                return self._create_empty_excel_report(exact_date, category_name, subcategory_name, branch_id, warehouse_id)

            print(f"Found {len(filtered_transaction_ids)} transactions to analyze")

            # Step 6: Build optimized query using the filtered transaction IDs
            from sqlalchemy import func, case
            query = db.session.query(
                FixedAsset.id.label('asset_id'),
                FixedAsset.name_ar.label('asset_name_ar'),
                FixedAsset.name_en.label('asset_name_en'),
                FixedAsset.product_code.label('product_code'),
                Category.category.label('category'),
                Category.subcategory.label('subcategory'),
                # Sum quantities for IN transactions (transaction_type = True)
                func.coalesce(
                    func.sum(
                        case(
                            (Transaction.transaction_type == True, AssetTransaction.quantity),
                            else_=0
                        )
                    ), 0
                ).label('total_quantity_in'),
                # Sum quantities for OUT transactions (transaction_type = False)
                func.coalesce(
                    func.sum(
                        case(
                            (Transaction.transaction_type == False, AssetTransaction.quantity),
                            else_=0
                        )
                    ), 0
                ).label('total_quantity_out'),
                # Sum amounts for IN transactions
                func.coalesce(
                    func.sum(
                        case(
                            (Transaction.transaction_type == True, AssetTransaction.amount),
                            else_=0
                        )
                    ), 0
                ).label('total_amount_in'),
                # Sum amounts for OUT transactions
                func.coalesce(
                    func.sum(
                        case(
                            (Transaction.transaction_type == False, AssetTransaction.amount),
                            else_=0
                        )
                    ), 0
                ).label('total_amount_out'),
                # Sum total costs for IN transactions (quantity × amount)
                func.coalesce(
                    func.sum(
                        case(
                            (Transaction.transaction_type == True, AssetTransaction.quantity * AssetTransaction.amount),
                            else_=0
                        )
                    ), 0
                ).label('total_cost_in'),
                # Sum total costs for OUT transactions
                func.coalesce(
                    func.sum(
                        case(
                            (Transaction.transaction_type == False, AssetTransaction.quantity * AssetTransaction.amount),
                            else_=0
                        )
                    ), 0
                ).label('total_cost_out')
            ).select_from(
                AssetTransaction
            ).join(
                Transaction, AssetTransaction.transaction_id == Transaction.id
            ).join(
                FixedAsset, AssetTransaction.asset_id == FixedAsset.id
            ).join(
                Category, FixedAsset.category_id == Category.id
            ).filter(
                # Only include asset transactions from our filtered transactions
                AssetTransaction.transaction_id.in_(filtered_transaction_ids)
            )

            # Step 7: Apply category and subcategory filters if provided
            if category_name:
                print(f"Filtering by category: {category_name}")
                query = query.filter(Category.category == category_name)
            
            if subcategory_name:
                print(f"Filtering by subcategory: {subcategory_name}")
                query = query.filter(Category.subcategory == subcategory_name)

            # Step 8: Group by asset to get asset-level aggregations
            query = query.group_by(
                FixedAsset.id,
                FixedAsset.name_ar,
                FixedAsset.name_en,
                FixedAsset.product_code,
                Category.category,
                Category.subcategory
            )

            # Step 9: Execute the optimized query
            print("Executing final aggregation query...")
            results = query.all()
            print(f"Found {len(results)} assets with transactions")

            # Step 10: Process results and create Excel file
            return self._create_excel_report(
                results, exact_date, category_name, subcategory_name, 
                branch_id, warehouse_id, len(filtered_transaction_ids)
            )

        except Exception as e:
            print(f"Excel report generation error: {str(e)}")
            return create_error_response(f"Excel report generation failed: {str(e)}", 500)

    def _create_excel_report(self, results, exact_date, category_name, subcategory_name, branch_id, warehouse_id, total_transactions):
        """Create Excel file with filter information and data"""
        try:
            # Create BytesIO buffer
            output = BytesIO()
            
            # Create workbook directly without pandas ExcelWriter initially
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Transaction Report"
            
            # Set column widths
            worksheet.column_dimensions['A'].width = 20
            worksheet.column_dimensions['B'].width = 25
            worksheet.column_dimensions['C'].width = 25
            worksheet.column_dimensions['D'].width = 15
            worksheet.column_dimensions['E'].width = 15
            worksheet.column_dimensions['F'].width = 15
            worksheet.column_dimensions['G'].width = 12
            worksheet.column_dimensions['H'].width = 12
            worksheet.column_dimensions['I'].width = 15
            worksheet.column_dimensions['J'].width = 15
            worksheet.column_dimensions['K'].width = 15
            worksheet.column_dimensions['L'].width = 15
            worksheet.column_dimensions['M'].width = 12
            worksheet.column_dimensions['N'].width = 15
            worksheet.column_dimensions['O'].width = 15
            
            current_row = 1
            
            # Add title
            worksheet.merge_cells(f'A{current_row}:O{current_row}')
            title_cell = worksheet[f'A{current_row}']
            title_cell.value = "Transaction Report"
            title_cell.font = openpyxl.styles.Font(size=16, bold=True)
            title_cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            current_row += 2
            
            # Add generation info
            worksheet[f'A{current_row}'] = "Generated At:"
            worksheet[f'B{current_row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            current_row += 1
            
            # Add filters section
            worksheet[f'A{current_row}'] = "Applied Filters:"
            worksheet[f'A{current_row}'].font = openpyxl.styles.Font(bold=True)
            current_row += 1
            
            worksheet[f'A{current_row}'] = "Date:"
            worksheet[f'B{current_row}'] = exact_date
            current_row += 1
            
            if category_name:
                worksheet[f'A{current_row}'] = "Category:"
                worksheet[f'B{current_row}'] = category_name
                current_row += 1
            
            if subcategory_name:
                worksheet[f'A{current_row}'] = "Subcategory:"
                worksheet[f'B{current_row}'] = subcategory_name
                current_row += 1
            
            if branch_id:
                worksheet[f'A{current_row}'] = "Branch ID:"
                worksheet[f'B{current_row}'] = branch_id
                current_row += 1
            
            if warehouse_id:
                worksheet[f'A{current_row}'] = "Warehouse ID:"
                worksheet[f'B{current_row}'] = warehouse_id
                current_row += 1
            
            worksheet[f'A{current_row}'] = "Total Transactions:"
            worksheet[f'B{current_row}'] = total_transactions
            current_row += 2
            
            # Process results into data
            asset_data = []
            total_summary = {
                'total_quantity_in': 0,
                'total_quantity_out': 0,
                'total_amount_in': 0,
                'total_amount_out': 0,
                'total_cost_in': 0,
                'total_cost_out': 0
            }
            
            for result in results:
                row_data = {
                    'Asset ID': result.asset_id,
                    'Asset Name (AR)': result.asset_name_ar,
                    'Asset Name (EN)': result.asset_name_en,
                    'Product Code': result.product_code,
                    'Category': result.category,
                    'Subcategory': result.subcategory,
                    'Qty IN': int(result.total_quantity_in),
                    'Qty OUT': int(result.total_quantity_out),
                    'Amount IN': float(result.total_amount_in),
                    'Amount OUT': float(result.total_amount_out),
                    'Cost IN': float(result.total_cost_in),
                    'Cost OUT': float(result.total_cost_out),
                    'Net Qty': int(result.total_quantity_in - result.total_quantity_out),
                    'Net Amount': float(result.total_amount_in - result.total_amount_out),
                    'Net Cost': float(result.total_cost_in - result.total_cost_out)
                }
                
                asset_data.append(row_data)
                
                # Add to totals
                total_summary['total_quantity_in'] += row_data['Qty IN']
                total_summary['total_quantity_out'] += row_data['Qty OUT']
                total_summary['total_amount_in'] += row_data['Amount IN']
                total_summary['total_amount_out'] += row_data['Amount OUT']
                total_summary['total_cost_in'] += row_data['Cost IN']
                total_summary['total_cost_out'] += row_data['Cost OUT']
            
            # Add data table
            if asset_data:
                # Add headers
                headers = ['Asset ID', 'Asset Name (AR)', 'Asset Name (EN)', 'Product Code', 'Category', 'Subcategory',
                          'Qty IN', 'Qty OUT', 'Amount IN', 'Amount OUT', 'Cost IN', 'Cost OUT', 
                          'Net Qty', 'Net Amount', 'Net Cost']
                
                header_row = current_row
                for col_num, header in enumerate(headers, 1):
                    cell = worksheet.cell(row=header_row, column=col_num)
                    cell.value = header
                    cell.font = openpyxl.styles.Font(bold=True)
                    cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                
                current_row += 1
                
                # Add data rows
                for row_data in asset_data:
                    for col_num, header in enumerate(headers, 1):
                        cell = worksheet.cell(row=current_row, column=col_num)
                        cell.value = row_data[header]
                    current_row += 1
                
                # Add totals row
                totals_row = current_row
                worksheet[f'A{totals_row}'] = "TOTALS"
                worksheet[f'A{totals_row}'].font = openpyxl.styles.Font(bold=True)
                
                # Calculate net totals
                net_quantity = total_summary['total_quantity_in'] - total_summary['total_quantity_out']
                net_amount = total_summary['total_amount_in'] - total_summary['total_amount_out']
                net_cost = total_summary['total_cost_in'] - total_summary['total_cost_out']
                
                worksheet[f'G{totals_row}'] = total_summary['total_quantity_in']
                worksheet[f'H{totals_row}'] = total_summary['total_quantity_out']
                worksheet[f'I{totals_row}'] = total_summary['total_amount_in']
                worksheet[f'J{totals_row}'] = total_summary['total_amount_out']
                worksheet[f'K{totals_row}'] = total_summary['total_cost_in']
                worksheet[f'L{totals_row}'] = total_summary['total_cost_out']
                worksheet[f'M{totals_row}'] = net_quantity
                worksheet[f'N{totals_row}'] = net_amount
                worksheet[f'O{totals_row}'] = net_cost
                
                # Style totals row
                for col in range(1, 16):  # A to O
                    cell = worksheet.cell(row=totals_row, column=col)
                    cell.font = openpyxl.styles.Font(bold=True)
                    cell.fill = openpyxl.styles.PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            
            else:
                # No data found
                worksheet[f'A{current_row}'] = "No transactions found for the specified filters."
                worksheet[f'A{current_row}'].font = openpyxl.styles.Font(italic=True)
            
            # Save workbook to BytesIO
            workbook.save(output)
            output.seek(0)
            
            # Generate filename
            filename = f"transaction_report_{exact_date.replace('-', '_')}.xlsx"
            
            return send_file(
                output,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
        except Exception as e:
            print(f"Error creating Excel file: {str(e)}")
            return create_error_response(f"Failed to create Excel file: {str(e)}", 500)
    
    def _create_empty_excel_report(self, exact_date, category_name, subcategory_name, branch_id, warehouse_id):
        """Create empty Excel file when no data is found"""
        try:
            output = BytesIO()
            
            # Create workbook directly
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Transaction Report"
            
            current_row = 1
            
            # Add title
            worksheet[f'A{current_row}'] = "Transaction Report"
            worksheet[f'A{current_row}'].font = openpyxl.styles.Font(size=16, bold=True)
            current_row += 2
            
            # Add generation info
            worksheet[f'A{current_row}'] = "Generated At:"
            worksheet[f'B{current_row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            current_row += 1
            
            # Add filters
            worksheet[f'A{current_row}'] = "Applied Filters:"
            worksheet[f'A{current_row}'].font = openpyxl.styles.Font(bold=True)
            current_row += 1
            
            worksheet[f'A{current_row}'] = "Date:"
            worksheet[f'B{current_row}'] = exact_date
            current_row += 1
            
            if category_name:
                worksheet[f'A{current_row}'] = "Category:"
                worksheet[f'B{current_row}'] = category_name
                current_row += 1
            
            if subcategory_name:
                worksheet[f'A{current_row}'] = "Subcategory:"
                worksheet[f'B{current_row}'] = subcategory_name
                current_row += 1
            
            if branch_id:
                worksheet[f'A{current_row}'] = "Branch ID:"
                worksheet[f'B{current_row}'] = branch_id
                current_row += 1
            
            if warehouse_id:
                worksheet[f'A{current_row}'] = "Warehouse ID:"
                worksheet[f'B{current_row}'] = warehouse_id
                current_row += 1
            
            current_row += 1
            worksheet[f'A{current_row}'] = "No transactions found for the specified filters."
            worksheet[f'A{current_row}'].font = openpyxl.styles.Font(italic=True, color="FF0000")
            
            # Save workbook to BytesIO
            workbook.save(output)
            output.seek(0)
            filename = f"transaction_report_{exact_date.replace('-', '_')}_empty.xlsx"
            
            return send_file(
                output,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
        except Exception as e:
            return create_error_response(f"Failed to create empty Excel file: {str(e)}", 500)
