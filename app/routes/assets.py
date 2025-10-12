from sqlalchemy import or_
from sqlalchemy.exc import IntegrityError
from flask import Blueprint, request, jsonify, send_file
from flask_restx import Resource, fields
from marshmallow import ValidationError
from .. import db
from ..models import FixedAsset, Category
from ..schemas import FixedAssetSchema, CategorySchema
from flask_jwt_extended import jwt_required
from ..utils import check_permission, generate_barcode, generate_unique_product_code, create_error_response, create_validation_error_response
from ..swagger import assets_ns, categories_ns, add_standard_responses, api
from ..swagger_models import (
    asset_model, asset_input_model, category_model, category_input_model,
    pagination_model, error_model, success_model, barcode_model, asset_search_response_model
)
import pandas as pd
from io import BytesIO
import openpyxl
import openpyxl.styles
from datetime import datetime

bp = Blueprint("assets", __name__, url_prefix="/assets")
asset_schema = FixedAssetSchema()
assets_schema = FixedAssetSchema(many=True)
category_schema = CategorySchema()
categories_schema = CategorySchema(many=True)

# Define the models outside the class first
bulk_summary_model = api.model('BulkSummary', {
    'total_processed': fields.Integer(description='Total number of assets processed'),
    'successfully_added': fields.Integer(description='Number of assets successfully added'),
    'rejected': fields.Integer(description='Number of assets rejected'),
    'success_rate': fields.String(description='Success rate percentage')
})

rejected_asset_model = api.model('RejectedAsset', {
    'asset_data': fields.Raw(description='Original asset data that was rejected'),
    'asset_name': fields.String(description='Asset name (if available)'),
    'errors': fields.List(fields.String, description='List of error messages')
})

bulk_create_result_model = api.model('BulkCreateResult', {
    'summary': fields.Nested(bulk_summary_model),
    'added_assets': fields.List(fields.Nested(asset_model), description='Successfully added assets'),
    'rejected_assets': fields.List(fields.Nested(rejected_asset_model), description='Rejected assets with error details')
})

# Add category bulk models after the existing bulk models
category_bulk_summary_model = api.model('CategoryBulkSummary', {
    'total_processed': fields.Integer(description='Total number of categories processed'),
    'successfully_added': fields.Integer(description='Number of categories successfully added'),
    'rejected': fields.Integer(description='Number of categories rejected'),
    'success_rate': fields.String(description='Success rate percentage')
})

rejected_category_model = api.model('RejectedCategory', {
    'category_data': fields.Raw(description='Original category data that was rejected'),
    'category_name': fields.String(description='Category name (if available)'),
    'errors': fields.List(fields.String, description='List of error messages')
})

category_bulk_create_result_model = api.model('CategoryBulkCreateResult', {
    'summary': fields.Nested(category_bulk_summary_model),
    'added_categories': fields.List(fields.Nested(category_model), description='Successfully added categories'),
    'rejected_categories': fields.List(fields.Nested(rejected_category_model), description='Rejected categories with error details')
})

# Add bulk update models after the existing bulk models
bulk_update_result_model = api.model('BulkUpdateResult', {
    'summary': fields.Nested(bulk_summary_model),
    'updated_assets': fields.List(fields.Nested(asset_model), description='Successfully updated assets'),
    'rejected_assets': fields.List(fields.Nested(rejected_asset_model), description='Rejected assets with error details')
})

@categories_ns.route("/")
class CategoryList(Resource):
    @categories_ns.doc('list_categories', security='Bearer Auth')
    @categories_ns.marshal_with(pagination_model)
    @categories_ns.param('page', 'Page number', type=int, default=1)
    @categories_ns.param('per_page', 'Items per page', type=int, default=10)
    @categories_ns.param('search', 'Search in category or subcategory names', type=str)
    @categories_ns.param('subcategory', 'Filter by subcategory name', type=str)
    @categories_ns.response(401, 'Unauthorized', error_model)
    @categories_ns.response(403, 'Forbidden', error_model)
    @jwt_required()
    def get(self):
        """Get all categories with pagination and optional search/filtering
        
        - search: Searches in both category and subcategory fields
        - subcategory: Filters by specific subcategory name
        """
        error = check_permission("can_read_asset")
        if error:
            return error

        page = request.args.get("page", 1, type=int)
        per_page = request.args.get("per_page", 10, type=int)
        search = request.args.get("search", "").strip()
        subcategory_filter = request.args.get("subcategory", "").strip()

        query = Category.query
        
        # Apply search filter if provided (searches in both category and subcategory)
        if search:
            search_pattern = f"%{search}%"
            query = query.filter(
                or_(
                    Category.category.ilike(search_pattern),
                    Category.subcategory.ilike(search_pattern)
                )
            )
        
        # Apply subcategory filter if provided
        if subcategory_filter:
            query = query.filter(Category.subcategory.ilike(f"%{subcategory_filter}%"))

        # Order results by ID descending for consistent ordering
        query = query.order_by(Category.id.desc())

        paginated = query.paginate(page=page, per_page=per_page, error_out=False)
        return {
            "items": categories_schema.dump(paginated.items),
            "total": paginated.total,
            "page": paginated.page,
            "pages": paginated.pages
        }

    @categories_ns.doc('create_category', security='Bearer Auth')
    @categories_ns.expect(category_input_model)
    @categories_ns.marshal_with(category_model, code=201)
    @categories_ns.response(400, 'Validation Error', error_model)
    @categories_ns.response(401, 'Unauthorized', error_model)
    @categories_ns.response(403, 'Forbidden', error_model)
    @categories_ns.response(500, 'Internal Server Error', error_model)
    @jwt_required()
    def post(self):
        """Create a new category"""
        error = check_permission("can_edit_asset")
        if error:
            return error

        try:
            data = category_schema.load(request.get_json())
            new_category = Category(**data)
            db.session.add(new_category)
            db.session.commit()
            return category_schema.dump(new_category), 201
        except ValidationError as err:
            return create_validation_error_response(err.messages)
        except IntegrityError as e:
            db.session.rollback()
            error_str = str(e.orig)
            
            if "duplicate key value violates unique constraint" in error_str:
                return create_error_response("A category with this information already exists in the system", 400)
            else:
                return create_error_response("The data provided violates database constraints", 400)
        except Exception as e:
            db.session.rollback()
            return create_error_response("An unexpected error occurred while creating the category", 500)


@categories_ns.route("/<int:category_id>")
class CategoryResource(Resource):
    @categories_ns.doc('get_category', security='Bearer Auth')
    @categories_ns.marshal_with(category_model)
    @categories_ns.response(401, 'Unauthorized', error_model)
    @categories_ns.response(403, 'Forbidden', error_model)
    @categories_ns.response(404, 'Category not found', error_model)
    @jwt_required()
    def get(self, category_id):
        """Get a specific category"""
        error = check_permission("can_read_asset")
        if error:
            return error

        category = db.session.get(Category, category_id)
        if not category:
            return create_error_response("Category not found", 404)
        return category_schema.dump(category)

    @categories_ns.doc('update_category', security='Bearer Auth')
    @categories_ns.expect(category_input_model)
    @categories_ns.marshal_with(category_model)
    @categories_ns.response(400, 'Validation Error', error_model)
    @categories_ns.response(401, 'Unauthorized', error_model)
    @categories_ns.response(403, 'Forbidden', error_model)
    @categories_ns.response(404, 'Category not found', error_model)
    @categories_ns.response(500, 'Internal Server Error', error_model)
    @jwt_required()
    def put(self, category_id):
        """Update a category"""
        error = check_permission("can_edit_asset")
        if error:
            return error

        category = db.session.get(Category, category_id)
        if not category:
            return create_error_response("Category not found", 404)

        try:
            data = category_schema.load(request.get_json(), partial=True)
            for key, value in data.items():
                setattr(category, key, value)
            db.session.commit()
            return category_schema.dump(category)
        except ValidationError as err:
            return create_validation_error_response(err.messages)
        except IntegrityError as e:
            db.session.rollback()
            error_str = str(e.orig)
            
            if "duplicate key value violates unique constraint" in error_str:
                return create_error_response("A category with this information already exists in the system", 400)
            else:
                return create_error_response("The data provided violates database constraints", 400)
        except Exception as e:
            db.session.rollback()
            return create_error_response("An unexpected error occurred while updating the category", 500)

    @categories_ns.doc('delete_category', security='Bearer Auth')
    @categories_ns.marshal_with(success_model)
    @categories_ns.response(400, 'Cannot delete category', error_model)
    @categories_ns.response(401, 'Unauthorized', error_model)
    @categories_ns.response(403, 'Forbidden', error_model)
    @categories_ns.response(404, 'Category not found', error_model)
    @jwt_required()
    def delete(self, category_id):
        """Delete a category"""
        error = check_permission("can_delete_asset")
        if error:
            return error

        category = db.session.get(Category, category_id)
        if not category:
            return create_error_response("Category not found", 404)

        try:
            db.session.delete(category)
            db.session.commit()
            return {"message": f"Category {category_id} deleted successfully"}
        except Exception as e:
            db.session.rollback()
            return create_error_response("Cannot delete category with associated assets", 400)

@assets_ns.route("/")
class AssetList(Resource):
    @assets_ns.doc('list_assets', security='Bearer Auth')
    @assets_ns.marshal_with(pagination_model)
    @assets_ns.param('page', 'Page number', type=int, default=1)
    @assets_ns.param('per_page', 'Items per page', type=int, default=10)
    @assets_ns.param('category_id', 'Filter assets by category ID', type=int)
    @assets_ns.param('subcategory', 'Filter assets by subcategory name', type=str)
    @assets_ns.response(401, 'Unauthorized', error_model)
    @assets_ns.response(403, 'Forbidden', error_model)
    # @jwt_required()
    def get(self):
        """Get all fixed assets with pagination and optional category/subcategory filtering
        
        - category_id: Filter by specific category ID
        - subcategory: Filter by subcategory name (case-insensitive partial match)
        """
        # error = check_permission("can_read_asset")
        # if error:
        #     return error

        page = request.args.get("page", 1, type=int)
        per_page = request.args.get("per_page", 10, type=int)
        category_id = request.args.get("category_id", type=int)
        subcategory = request.args.get("subcategory", "").strip()

        query = FixedAsset.query
        
        # Filter by category_id if provided
        if category_id:
            query = query.filter_by(category_id=category_id)
        
        # Filter by subcategory if provided
        if subcategory:
            # Join with Category table to filter by subcategory name
            query = query.join(Category).filter(Category.subcategory.ilike(f"%{subcategory}%"))

        # Order by ID descending for consistent ordering
        query = query.order_by(FixedAsset.id.desc())

        paginated = query.paginate(page=page, per_page=per_page)
        return {
            "items": assets_schema.dump(paginated.items),
            "total": paginated.total,
            "page": paginated.page,
            "pages": paginated.pages
        }
    
    @assets_ns.doc('create_asset', security='Bearer Auth')
    @assets_ns.expect(asset_input_model)
    @assets_ns.marshal_with(asset_model, code=201)
    @assets_ns.response(400, 'Validation Error', error_model)
    @assets_ns.response(401, 'Unauthorized', error_model)
    @assets_ns.response(403, 'Forbidden', error_model)
    @assets_ns.response(500, 'Internal Server Error', error_model)
    @jwt_required()
    def post(self):
        """Create a new fixed asset"""
        error = check_permission("can_edit_asset")
        if error:
            return error

        try:
            data = asset_schema.load(request.get_json())
            new_asset = FixedAsset(**data)
            db.session.add(new_asset)
            db.session.commit()
            return asset_schema.dump(new_asset), 201
        except ValidationError as err:
            return create_validation_error_response(err.messages)
        except IntegrityError as e:
            db.session.rollback()
            error_str = str(e.orig)
            
            if "duplicate key value violates unique constraint" in error_str and "product_code" in error_str:
                return create_error_response("The product code you provided is already in use. Please use a different product code", 400, "product_code")
            elif "duplicate key value violates unique constraint" in error_str:
                return create_error_response("A record with this information already exists in the system", 400)
            else:
                return create_error_response("The data provided violates database constraints", 400)
        except Exception as e:
            db.session.rollback()
            return create_error_response("An unexpected error occurred while creating the asset", 500)


@assets_ns.route("/<int:asset_id>/barcode")
class AssetBarcode(Resource):
    @assets_ns.doc('get_asset_barcode', security='Bearer Auth')
    @assets_ns.marshal_with(barcode_model, code=200, description='Successfully generated barcode')
    @assets_ns.response(401, 'Unauthorized', error_model)
    @assets_ns.response(403, 'Forbidden', error_model)
    @assets_ns.response(404, 'Asset not found', error_model)
    @jwt_required()
    def get(self, asset_id):
        """Generate a barcode for a specific asset"""
        error = check_permission("can_print_barcode")
        if error:
            return error
            
        # Get the asset
        asset = db.session.query(FixedAsset).filter_by(id=asset_id).first()
        if not asset:
            return create_error_response("Asset not found", 404)
            
        # Check if asset has a product code, if not generate one
        if not asset.product_code:
            asset.product_code = generate_unique_product_code()
            db.session.commit()
            
        # Generate barcode
        barcode_data = generate_barcode(asset.product_code)
        return barcode_data



@assets_ns.route("/<int:asset_id>")
class AssetResource(Resource):
    @assets_ns.doc('get_asset', security='Bearer Auth')
    @assets_ns.marshal_with(asset_model)
    @assets_ns.response(401, 'Unauthorized', error_model)
    @assets_ns.response(403, 'Forbidden', error_model)
    @assets_ns.response(404, 'Asset not found', error_model)
    @jwt_required()
    def get(self, asset_id):
        """Get a specific asset"""
        error = check_permission("can_read_asset")
        if error:
            return error

        asset = db.session.get(FixedAsset, asset_id)
        if not asset:
            return create_error_response("Asset not found", 404)
        return asset_schema.dump(asset)

    @assets_ns.doc('update_asset', security='Bearer Auth')
    @assets_ns.expect(asset_input_model)
    @assets_ns.marshal_with(asset_model)
    @assets_ns.response(400, 'Validation Error', error_model)
    @assets_ns.response(401, 'Unauthorized', error_model)
    @assets_ns.response(403, 'Forbidden', error_model)
    @assets_ns.response(404, 'Asset not found', error_model)
    @assets_ns.response(500, 'Internal Server Error', error_model)
    @jwt_required()
    def put(self, asset_id):
        """Update a specific asset"""
        error = check_permission("can_edit_asset")
        if error:
            return error

        asset = db.session.get(FixedAsset, asset_id)
        if not asset:
            return create_error_response("Asset not found", 404)

        try:
            data = asset_schema.load(request.get_json(), partial=True)
            for key, value in data.items():
                setattr(asset, key, value)
            db.session.commit()
            return asset_schema.dump(asset)
        except ValidationError as err:
            return create_validation_error_response(err.messages)
        except IntegrityError as e:
            db.session.rollback()
            error_str = str(e.orig)
            
            if "duplicate key value violates unique constraint" in error_str and "product_code" in error_str:
                return create_error_response("The product code you provided is already in use by another asset. Please use a different product code", 400, "product_code")
            elif "duplicate key value violates unique constraint" in error_str:
                return create_error_response("A record with this information already exists in the system", 400)
            else:
                return create_error_response("The data provided violates database constraints", 400)
        except Exception as e:
            db.session.rollback()
            return create_error_response("An unexpected error occurred while updating the asset", 500)

    @assets_ns.doc('delete_asset', security='Bearer Auth')
    @assets_ns.marshal_with(success_model)
    @assets_ns.response(401, 'Unauthorized', error_model)
    @assets_ns.response(403, 'Forbidden', error_model)
    @assets_ns.response(404, 'Asset not found', error_model)
    @jwt_required()
    def delete(self, asset_id):
        """Delete a specific asset"""
        error = check_permission("can_delete_asset")
        if error:
            return error

        asset = db.session.get(FixedAsset, asset_id)
        if not asset:
            return create_error_response("Asset not found", 404)

        db.session.delete(asset)
        db.session.commit()
        return {"message": f"Asset {asset_id} deleted successfully"}



# Searching Endpoint
@assets_ns.route("/search")
class AssetSearch(Resource):
    @assets_ns.doc('search_assets', security='Bearer Auth')
    @assets_ns.param('q', 'Search query (text for name search or number for barcode search)', required=True, type=str)
    @assets_ns.param('page', 'Page number', type=int, default=1)
    @assets_ns.param('per_page', 'Items per page', type=int, default=10)
    @assets_ns.marshal_with(asset_search_response_model)
    @assets_ns.response(400, 'Missing Search Query', error_model)
    @assets_ns.response(401, 'Unauthorized', error_model)
    @assets_ns.response(403, 'Forbidden', error_model)
    @assets_ns.response(500, 'Search Error', error_model)
    @jwt_required()
    def get(self):
        """Search assets by name (text) or product code/barcode (number)
        
        - If query contains letters: searches in name_ar and name_en fields
        - If query is numeric: searches by exact product_code match
        """
        error = check_permission("can_read_asset")
        if error:
            return error

        search_query = request.args.get('q', '').strip()
        if not search_query:
            return create_error_response("Search query parameter 'q' is required", 400)

        page = request.args.get("page", 1, type=int)
        per_page = request.args.get("per_page", 10, type=int)

        try:
            # Build base query
            query = FixedAsset.query.filter(FixedAsset.is_active == True)
            
            
            # Determine search type and build search conditions
            search_conditions = []
            
            # Check if query is purely numeric (for barcode search)
            if search_query.isdigit():
                # Pure number - search by product_code (exact match)
                search_conditions.append(FixedAsset.product_code == search_query)
            else:
                # Contains letters - search by name (partial match, case-insensitive)
                name_search = f"%{search_query}%"
                search_conditions.extend([
                    FixedAsset.name_ar.ilike(name_search),
                    FixedAsset.name_en.ilike(name_search)
                ])
                
                # Also check if it might be a product_code (partial match)
                if search_query.replace('-', '').replace('_', '').isalnum():
                    search_conditions.append(FixedAsset.product_code.ilike(f"%{search_query}%"))
            
            # Apply search conditions with OR logic
            if search_conditions:
                query = query.filter(or_(*search_conditions))
            else:
                # If no conditions, return empty result
                return {
                    "items": [],
                    "total": 0,
                    "page": page,
                    "pages": 0
                }
            
            # Order results by ID descending for consistent ordering
            query = query.order_by(FixedAsset.id.desc())
            
            # Execute paginated query
            paginated = query.paginate(page=page, per_page=per_page, error_out=False)
            
            return {
                "items": assets_schema.dump(paginated.items),
                "total": paginated.total,
                "page": paginated.page,
                "pages": paginated.pages
            }
            
        except Exception as e:
            print(f"Search error: {str(e)}")
            return create_error_response(f"Search error: {str(e)}", 500)

@assets_ns.route("/bulk")
class AssetBulkCreate(Resource):
    @assets_ns.doc('bulk_create_assets', security='Bearer Auth')
    @assets_ns.expect([api.model('AssetBulkInput', {
        'name_ar': fields.String(required=True, description='Arabic name'),
        'name_en': fields.String(required=True, description='English name'),
        'product_code': fields.String(description='Product code'),
        'category': fields.String(required=True, description='Category name (will be mapped to category_id)'),
        'is_active': fields.Boolean(description='Active status (default: true)')
    })], description='List of assets to create')
    @assets_ns.response(200, 'Bulk operation completed', bulk_create_result_model)
    @assets_ns.response(400, 'Invalid input data', error_model)
    @assets_ns.response(401, 'Unauthorized', error_model)
    @assets_ns.response(403, 'Forbidden', error_model)
    @jwt_required()
    def post(self):
        """Bulk create multiple assets with category name mapping and automatic quantity setting
        
        First validates all assets, then adds only the valid ones to the database.
        
        Key features:
        - Maps category names to category_id automatically
        - Forces quantity to always be 0 regardless of input
        - Provides detailed error reporting
        
        Returns a comprehensive report including:
        - Summary statistics (total, added, rejected, success rate)
        - List of successfully added assets
        - List of rejected assets with detailed error information
        """
        error = check_permission("can_edit_asset")
        if error:
            return error

        try:
            # Get the request data
            request_data = request.get_json()
            
            # Validate that we received a list
            if not isinstance(request_data, list):
                return create_error_response("Request body must be a list of assets", 400)
            
            if not request_data:
                return create_error_response("Asset list cannot be empty", 400)
            
            # Initialize lists for validation phase
            total_processed = len(request_data)
            valid_assets = []
            rejected_assets = []
            
            # Cache for category lookups to avoid repeated database queries
            category_cache = {}
            
            # Phase 1: Validate all assets without committing to database
            for index, asset_data in enumerate(request_data):
                # Get asset name for reporting
                asset_name = (
                    asset_data.get('name_ar') or 
                    asset_data.get('name_en') or 
                    f"Asset #{index + 1}"
                )
                
                try:
                    # Force quantity to always be 0
                    asset_data_modified = asset_data.copy()
                    asset_data_modified['quantity'] = 0
                    
                    # Handle category name to category_id mapping
                    category_name = asset_data.get('category')
                    if not category_name:
                        rejected_assets.append({
                            'asset_data': asset_data,
                            'asset_name': asset_name,
                            'errors': ['Category name is required']
                        })
                        continue
                    
                    # Look up category_id from category name (with caching)
                    if category_name not in category_cache:
                        category = Category.query.filter_by(category=category_name).first()
                        if category:
                            category_cache[category_name] = category.id
                        else:
                            category_cache[category_name] = None
                    
                    category_id = category_cache[category_name]
                    if category_id is None:
                        rejected_assets.append({
                            'asset_data': asset_data,
                            'asset_name': asset_name,
                            'errors': [f"Category '{category_name}' not found in database"]
                        })
                        continue
                    
                    # Replace category name with category_id for validation
                    asset_data_modified['category_id'] = category_id
                    del asset_data_modified['category']  # Remove category name from data
                    
                    # Validate the asset data using schema
                    validated_data = asset_schema.load(asset_data_modified)
                    
                    # Check for duplicate product_code in current batch
                    if validated_data.get('product_code'):
                        existing_in_batch = any(
                            asset.get('product_code') == validated_data['product_code'] 
                            for asset in valid_assets
                        )
                        if existing_in_batch:
                            rejected_assets.append({
                                'asset_data': asset_data,
                                'asset_name': asset_name,
                                'errors': [f"Product code '{validated_data['product_code']}' is duplicated in this batch"]
                            })
                            continue
                    
                    # Check for duplicate names in current batch
                    name_ar_duplicate = any(
                        asset.get('name_ar') == validated_data['name_ar'] 
                        for asset in valid_assets
                    )
                    name_en_duplicate = any(
                        asset.get('name_en') == validated_data['name_en'] 
                        for asset in valid_assets
                    )
                    
                    if name_ar_duplicate:
                        rejected_assets.append({
                            'asset_data': asset_data,
                            'asset_name': asset_name,
                            'errors': [f"Arabic name '{validated_data['name_ar']}' is duplicated in this batch"]
                        })
                        continue
                    
                    if name_en_duplicate:
                        rejected_assets.append({
                            'asset_data': asset_data,
                            'asset_name': asset_name,
                            'errors': [f"English name '{validated_data['name_en']}' is duplicated in this batch"]
                        })
                        continue
                    
                    # Check for existing records in database
                    existing_product_code = None
                    if validated_data.get('product_code'):
                        existing_product_code = FixedAsset.query.filter_by(product_code=validated_data['product_code']).first()
                    
                    existing_name_ar = FixedAsset.query.filter_by(name_ar=validated_data['name_ar']).first()
                    existing_name_en = FixedAsset.query.filter_by(name_en=validated_data['name_en']).first()
                    
                    if existing_product_code:
                        rejected_assets.append({
                            'asset_data': asset_data,
                            'asset_name': asset_name,
                            'errors': [f"Product code '{validated_data['product_code']}' already exists in database"]
                        })
                        continue
                    
                    if existing_name_ar:
                        rejected_assets.append({
                            'asset_data': asset_data,
                            'asset_name': asset_name,
                            'errors': [f"Arabic name '{validated_data['name_ar']}' already exists in database"]
                        })
                        continue
                    
                    if existing_name_en:
                        rejected_assets.append({
                            'asset_data': asset_data,
                            'asset_name': asset_name,
                            'errors': [f"English name '{validated_data['name_en']}' already exists in database"]
                        })
                        continue
                    
                    # If we get here, the asset is valid
                    valid_assets.append({
                        'data': validated_data,
                        'original': asset_data,
                        'name': asset_name
                    })
                    
                except ValidationError as ve:
                    # Schema validation error
                    error_messages = []
                    for field, messages in ve.messages.items():
                        if isinstance(messages, list):
                            error_messages.extend([f"{field}: {msg}" for msg in messages])
                        else:
                            error_messages.append(f"{field}: {messages}")
                    
                    rejected_assets.append({
                        'asset_data': asset_data,
                        'asset_name': asset_name,
                        'errors': error_messages
                    })
                
                except Exception as e:
                    rejected_assets.append({
                        'asset_data': asset_data,
                        'asset_name': asset_name,
                        'errors': [f"Validation error: {str(e)}"]
                    })
            
            # Phase 2: Add all valid assets to database
            successfully_added = []
            if valid_assets:
                try:
                    for asset_info in valid_assets:
                        new_asset = FixedAsset(**asset_info['data'])
                        db.session.add(new_asset)
                    
                    # Commit all at once
                    db.session.commit()
                    
                    # Refresh objects to get IDs and relationships
                    for asset_info in valid_assets:
                        # Find the corresponding committed asset
                        committed_asset = FixedAsset.query.filter_by(
                            name_ar=asset_info['data']['name_ar'],
                            name_en=asset_info['data']['name_en']
                        ).first()
                        if committed_asset:
                            successfully_added.append(committed_asset)
                
                except Exception as e:
                    db.session.rollback()
                    # If database commit fails, move all valid assets to rejected
                    for asset_info in valid_assets:
                        rejected_assets.append({
                            'asset_data': asset_info['original'],
                            'asset_name': asset_info['name'],
                            'errors': [f"Database error during commit: {str(e)}"]
                        })
                    successfully_added = []
            
            # Calculate statistics
            added_count = len(successfully_added)
            rejected_count = len(rejected_assets)
            success_rate = f"{(added_count / total_processed * 100):.1f}%" if total_processed > 0 else "0%"
            
            # Prepare response
            result = {
                'summary': {
                    'total_processed': total_processed,
                    'successfully_added': added_count,
                    'rejected': rejected_count,
                    'success_rate': success_rate
                },
                'added_assets': assets_schema.dump(successfully_added),
                'rejected_assets': rejected_assets
            }
            
            return result, 200
            
        except Exception as e:
            db.session.rollback()
            return create_error_response(f"Bulk operation failed: {str(e)}", 500)

@categories_ns.route("/bulk")
class CategoryBulkCreate(Resource):
    @categories_ns.doc('bulk_create_categories', security='Bearer Auth')
    @categories_ns.expect([category_input_model], description='List of categories to create')
    @categories_ns.response(200, 'Bulk operation completed', category_bulk_create_result_model)
    @categories_ns.response(400, 'Invalid input data', error_model)
    @categories_ns.response(401, 'Unauthorized', error_model)
    @categories_ns.response(403, 'Forbidden', error_model)
    # @jwt_required()
    def post(self):
        """Bulk create multiple categories with detailed error reporting
        
        First validates all categories, then adds only the valid ones to the database.
        
        Returns a comprehensive report including:
        - Summary statistics (total, added, rejected, success rate)
        - List of successfully added categories
        - List of rejected categories with detailed error information
        """
        # error = check_permission("can_edit_asset")
        # if error:
        #     return error

        try:
            # Get the request data
            request_data = request.get_json()
            
            # Validate that we received a list
            if not isinstance(request_data, list):
                return create_error_response("Request body must be a list of categories", 400)
            
            if not request_data:
                return create_error_response("Category list cannot be empty", 400)
            
            # Initialize lists for validation phase
            total_processed = len(request_data)
            valid_categories = []
            rejected_categories = []
            
            # Phase 1: Validate all categories without committing to database
            for index, category_data in enumerate(request_data):
                # Get category name for reporting
                category_name = (
                    category_data.get('category') or 
                    f"Category #{index + 1}"
                )
                
                try:
                    # Validate the category data using schema
                    validated_data = category_schema.load(category_data)
                    
                    # Check for duplicate category in current batch
                    # Fix: Access the nested data structure correctly
                    existing_in_batch = any(
                        cat['data'].get('category') == validated_data['category'] 
                        for cat in valid_categories
                    )
                    if existing_in_batch:
                        rejected_categories.append({
                            'category_data': category_data,
                            'category_name': category_name,
                            'errors': [f"Category '{validated_data['category']}' is duplicated in this batch"]
                        })
                        continue
                    
                    # Check for existing category in database (only category name matters for uniqueness)
                    existing_category = Category.query.filter_by(category=validated_data['category']).first()
                    if existing_category:
                        rejected_categories.append({
                            'category_data': category_data,
                            'category_name': category_name,
                            'errors': [f"Category '{validated_data['category']}' already exists in database"]
                        })
                        continue
                    
                    # If we get here, the category is valid
                    valid_categories.append({
                        'data': validated_data,
                        'original': category_data,
                        'name': category_name
                    })
                    
                except ValidationError as ve:
                    # Schema validation error
                    error_messages = []
                    for field, messages in ve.messages.items():
                        if isinstance(messages, list):
                            error_messages.extend([f"{field}: {msg}" for msg in messages])
                        else:
                            error_messages.append(f"{field}: {messages}")
                    
                    rejected_categories.append({
                        'category_data': category_data,
                        'category_name': category_name,
                        'errors': error_messages
                    })
                
                except Exception as e:
                    rejected_categories.append({
                        'category_data': category_data,
                        'category_name': category_name,
                        'errors': [f"Validation error: {str(e)}"]
                    })
            
            # Phase 2: Add all valid categories to database
            successfully_added = []
            if valid_categories:
                try:
                    for category_info in valid_categories:
                        new_category = Category(**category_info['data'])
                        db.session.add(new_category)
                    
                    # Commit all at once
                    db.session.commit()
                    
                    # Refresh objects to get IDs
                    for category_info in valid_categories:
                        # Find the corresponding committed category
                        committed_category = Category.query.filter_by(
                            category=category_info['data']['category']
                        ).first()
                        if committed_category:
                            successfully_added.append(committed_category)
                
                except IntegrityError as ie:
                    db.session.rollback()
                    error_str = str(ie.orig).lower()
                    
                    # Handle specific database constraint violations
                    if "unique constraint failed: categories.category" in error_str:
                        # If we get here, it means our validation missed something
                        # Move all categories to rejected with specific error
                        for category_info in valid_categories:
                            rejected_categories.append({
                                'category_data': category_info['original'],
                                'category_name': category_info['name'],
                                'errors': [f"Category '{category_info['data']['category']}' already exists in database"]
                            })
                    else:
                        # Other integrity errors
                        for category_info in valid_categories:
                            rejected_categories.append({
                                'category_data': category_info['original'],
                                'category_name': category_info['name'],
                                'errors': [f"Database constraint violation: {str(ie.orig)}"]
                            })
                    successfully_added = []
                
                except Exception as e:
                    db.session.rollback()
                    # If database commit fails, move all valid categories to rejected
                    for category_info in valid_categories:
                        rejected_categories.append({
                            'category_data': category_info['original'],
                            'category_name': category_info['name'],
                            'errors': [f"Database error during commit: {str(e)}"]
                        })
                    successfully_added = []
            
            # Calculate statistics
            added_count = len(successfully_added)
            rejected_count = len(rejected_categories)
            success_rate = f"{(added_count / total_processed * 100):.1f}%" if total_processed > 0 else "0%"
            
            # Prepare response
            result = {
                'summary': {
                    'total_processed': total_processed,
                    'successfully_added': added_count,
                    'rejected': rejected_count,
                    'success_rate': success_rate
                },
                'added_categories': categories_schema.dump(successfully_added),
                'rejected_categories': rejected_categories
            }
            
            return result, 200
            
        except Exception as e:
            db.session.rollback()
            return create_error_response(f"Bulk operation failed: {str(e)}", 500)

@assets_ns.route("/bulk-update")
class AssetBulkUpdate(Resource):
    @assets_ns.doc('bulk_update_assets', security='Bearer Auth')
    @assets_ns.expect([api.model('AssetUpdateInput', {
        'id': fields.Integer(required=True, description='Asset ID to update'),
        'name_ar': fields.String(description='Arabic name'),
        'name_en': fields.String(description='English name'),
        'product_code': fields.String(description='Product code'),
        'category': fields.String(description='Category name (will be mapped to category_id)'),
        'is_active': fields.Boolean(description='Active status')
    })], description='List of assets to update')
    @assets_ns.response(200, 'Bulk update completed', bulk_update_result_model)
    @assets_ns.response(400, 'Invalid input data', error_model)
    @assets_ns.response(401, 'Unauthorized', error_model)
    @assets_ns.response(403, 'Forbidden', error_model)
    @jwt_required()
    def put(self):
        """Bulk update multiple assets with category name mapping and quantity preservation
        
        First validates all asset updates, then applies only the valid ones to the database.
        
        Key features:
        - Maps category names to category_id automatically
        - Preserves existing quantity (ignores any quantity input)
        - Provides detailed error reporting
        
        Returns a comprehensive report including:
        - Summary statistics (total, updated, rejected, success rate)
        - List of successfully updated assets
        - List of rejected assets with detailed error information
        """
        error = check_permission("can_edit_asset")
        if error:
            return error

        try:
            # Get the request data
            request_data = request.get_json()
            
            # Validate that we received a list
            if not isinstance(request_data, list):
                return create_error_response("Request body must be a list of asset updates", 400)
            
            if not request_data:
                return create_error_response("Asset update list cannot be empty", 400)
            
            # Initialize lists for validation phase
            total_processed = len(request_data)
            valid_updates = []
            rejected_assets = []
            
            # Cache for category lookups to avoid repeated database queries
            category_cache = {}
            
            # Phase 1: Validate all asset updates without committing to database
            for index, asset_data in enumerate(request_data):
                # Get asset identifier for reporting
                asset_id = asset_data.get('id')
                asset_name = (
                    asset_data.get('name_ar') or 
                    asset_data.get('name_en') or 
                    f"Asset ID {asset_id}" if asset_id else f"Asset #{index + 1}"
                )
                
                try:
                    # Check if asset ID is provided
                    if not asset_id:
                        rejected_assets.append({
                            'asset_data': asset_data,
                            'asset_name': asset_name,
                            'errors': ['Asset ID is required for update operation']
                        })
                        continue
                    
                    # Check if asset exists in database
                    existing_asset = db.session.get(FixedAsset, asset_id)
                    if not existing_asset:
                        rejected_assets.append({
                            'asset_data': asset_data,
                            'asset_name': asset_name,
                            'errors': [f'Asset with ID {asset_id} not found in database']
                        })
                        continue
                    
                    # Prepare update data - exclude ID and quantity
                    update_data = {k: v for k, v in asset_data.items() if k not in ['id', 'quantity']}
                    
                    # Handle category name to category_id mapping if category is provided
                    if 'category' in update_data:
                        category_name = update_data['category']
                        
                        # Look up category_id from category name (with caching)
                        if category_name not in category_cache:
                            category = Category.query.filter_by(category=category_name).first()
                            if category:
                                category_cache[category_name] = category.id
                            else:
                                category_cache[category_name] = None
                        
                        category_id = category_cache[category_name]
                        if category_id is None:
                            rejected_assets.append({
                                'asset_data': asset_data,
                                'asset_name': asset_name,
                                'errors': [f"Category '{category_name}' not found in database"]
                            })
                            continue
                        
                        # Replace category name with category_id
                        update_data['category_id'] = category_id
                        del update_data['category']
                    
                    # Validate the asset data using schema (partial=True for updates)
                    validated_data = asset_schema.load(update_data, partial=True)
                    
                    # Check for duplicate product_code in current batch (excluding current asset)
                    if validated_data.get('product_code'):
                        existing_in_batch = any(
                            update['validated_data'].get('product_code') == validated_data['product_code'] 
                            and update['asset_id'] != asset_id
                            for update in valid_updates
                        )
                        if existing_in_batch:
                            rejected_assets.append({
                                'asset_data': asset_data,
                                'asset_name': asset_name,
                                'errors': [f"Product code '{validated_data['product_code']}' is duplicated in this batch"]
                            })
                            continue
                    
                    # Check for duplicate names in current batch (excluding current asset)
                    if validated_data.get('name_ar'):
                        name_ar_duplicate = any(
                            update['validated_data'].get('name_ar') == validated_data['name_ar'] 
                            and update['asset_id'] != asset_id
                            for update in valid_updates
                        )
                        if name_ar_duplicate:
                            rejected_assets.append({
                                'asset_data': asset_data,
                                'asset_name': asset_name,
                                'errors': [f"Arabic name '{validated_data['name_ar']}' is duplicated in this batch"]
                            })
                            continue
                    
                    if validated_data.get('name_en'):
                        name_en_duplicate = any(
                            update['validated_data'].get('name_en') == validated_data['name_en'] 
                            and update['asset_id'] != asset_id
                            for update in valid_updates
                        )
                        if name_en_duplicate:
                            rejected_assets.append({
                                'asset_data': asset_data,
                                'asset_name': asset_name,
                                'errors': [f"English name '{validated_data['name_en']}' is duplicated in this batch"]
                            })
                            continue
                    
                    # Check if category exists in database (if category_id is being updated)
                    if validated_data.get('category_id'):
                        if not db.session.get(Category, validated_data['category_id']):
                            rejected_assets.append({
                                'asset_data': asset_data,
                                'asset_name': asset_name,
                                'errors': [f"Invalid category ID '{validated_data['category_id']}' - category does not exist"]
                            })
                            continue
                    
                    # Check for existing records in database (excluding current asset)
                    if validated_data.get('product_code'):
                        existing_product_code = FixedAsset.query.filter(
                            FixedAsset.product_code == validated_data['product_code'],
                            FixedAsset.id != asset_id
                        ).first()
                        if existing_product_code:
                            rejected_assets.append({
                                'asset_data': asset_data,
                                'asset_name': asset_name,
                                'errors': [f"Product code '{validated_data['product_code']}' already exists in database (used by asset ID {existing_product_code.id})"]
                            })
                            continue
                    
                    if validated_data.get('name_ar'):
                        existing_name_ar = FixedAsset.query.filter(
                            FixedAsset.name_ar == validated_data['name_ar'],
                            FixedAsset.id != asset_id
                        ).first()
                        if existing_name_ar:
                            rejected_assets.append({
                                'asset_data': asset_data,
                                'asset_name': asset_name,
                                'errors': [f"Arabic name '{validated_data['name_ar']}' already exists in database (used by asset ID {existing_name_ar.id})"]
                            })
                            continue
                    
                    if validated_data.get('name_en'):
                        existing_name_en = FixedAsset.query.filter(
                            FixedAsset.name_en == validated_data['name_en'],
                            FixedAsset.id != asset_id
                        ).first()
                        if existing_name_en:
                            rejected_assets.append({
                                'asset_data': asset_data,
                                'asset_name': asset_name,
                                'errors': [f"English name '{validated_data['name_en']}' already exists in database (used by asset ID {existing_name_en.id})"]
                            })
                            continue
                    
                    # If we get here, the update is valid
                    valid_updates.append({
                        'asset_id': asset_id,
                        'asset': existing_asset,
                        'validated_data': validated_data,
                        'original': asset_data,
                        'name': asset_name
                    })
                    
                except ValidationError as ve:
                    # Schema validation error
                    error_messages = []
                    for field, messages in ve.messages.items():
                        if isinstance(messages, list):
                            error_messages.extend([f"{field}: {msg}" for msg in messages])
                        else:
                            error_messages.append(f"{field}: {messages}")
                    
                    rejected_assets.append({
                        'asset_data': asset_data,
                        'asset_name': asset_name,
                        'errors': error_messages
                    })
                
                except Exception as e:
                    rejected_assets.append({
                        'asset_data': asset_data,
                        'asset_name': asset_name,
                        'errors': [f"Validation error: {str(e)}"]
                    })
            
            # Phase 2: Apply all valid updates to database
            successfully_updated = []
            if valid_updates:
                try:
                    for update_info in valid_updates:
                        asset = update_info['asset']
                        validated_data = update_info['validated_data']
                        
                        # Apply updates to the asset (quantity is preserved/not updated)
                        for key, value in validated_data.items():
                            setattr(asset, key, value)
                    
                    # Commit all changes at once
                    db.session.commit()
                    
                    # Add updated assets to success list
                    for update_info in valid_updates:
                        successfully_updated.append(update_info['asset'])
                
                except IntegrityError as ie:
                    db.session.rollback()
                    error_str = str(ie.orig).lower()
                    
                    # Handle specific database constraint violations
                    if "unique constraint failed" in error_str:
                        for update_info in valid_updates:
                            if "product_code" in error_str:
                                error_msg = f"Product code '{update_info['validated_data'].get('product_code', 'N/A')}' already exists in database"
                            elif "name_ar" in error_str:
                                error_msg = f"Arabic name '{update_info['validated_data'].get('name_ar', 'N/A')}' already exists in database"
                            elif "name_en" in error_str:
                                error_msg = f"English name '{update_info['validated_data'].get('name_en', 'N/A')}' already exists in database"
                            else:
                                error_msg = "Duplicate record - asset with this information already exists"
                            
                            rejected_assets.append({
                                'asset_data': update_info['original'],
                                'asset_name': update_info['name'],
                                'errors': [error_msg]
                            })
                    else:
                        # Other integrity errors
                        for update_info in valid_updates:
                            rejected_assets.append({
                                'asset_data': update_info['original'],
                                'asset_name': update_info['name'],
                                'errors': [f"Database constraint violation: {str(ie.orig)}"]
                            })
                    successfully_updated = []
                
                except Exception as e:
                    db.session.rollback()
                    # If database commit fails, move all valid updates to rejected
                    for update_info in valid_updates:
                        rejected_assets.append({
                            'asset_data': update_info['original'],
                            'asset_name': update_info['name'],
                            'errors': [f"Database error during commit: {str(e)}"]
                        })
                    successfully_updated = []
            
            # Calculate statistics
            updated_count = len(successfully_updated)
            rejected_count = len(rejected_assets)
            success_rate = f"{(updated_count / total_processed * 100):.1f}%" if total_processed > 0 else "0%"
            
            # Prepare response
            result = {
                'summary': {
                    'total_processed': total_processed,
                    'successfully_added': updated_count,  # Using 'successfully_added' to match the model
                    'rejected': rejected_count,
                    'success_rate': success_rate
                },
                'updated_assets': assets_schema.dump(successfully_updated),
                'rejected_assets': rejected_assets
            }
            
            return result, 200
            
        except Exception as e:
            db.session.rollback()
            return create_error_response(f"Bulk update operation failed: {str(e)}", 500)

@assets_ns.route("/export-excel")
class AssetExcelExport(Resource):
    @assets_ns.doc('export_assets_excel', security='Bearer Auth')
    @assets_ns.param('category_id', 'Filter assets by category ID', type=int)
    @assets_ns.param('subcategory', 'Filter assets by subcategory name', type=str)
    @assets_ns.response(200, 'Successfully generated Excel export')
    @assets_ns.response(401, 'Unauthorized', error_model)
    @assets_ns.response(403, 'Forbidden', error_model)
    @assets_ns.response(500, 'Internal Server Error', error_model)
    @jwt_required()
    def get(self):
        """Export all assets to Excel file with optional category/subcategory filtering
        
        Returns Excel file containing:
        - Filter information at the top (if filters are applied)
        - Complete list of assets with all details
        
        Optional filters:
        - category_id: Filter by specific category ID
        - subcategory: Filter by subcategory name (case-insensitive partial match)
        """
        error = check_permission("can_read_asset")
        if error:
            return error

        # Get filter parameters
        category_id = request.args.get("category_id", type=int)
        subcategory = request.args.get("subcategory", "").strip()

        try:
            # Build query without relationship loading since it doesn't exist
            query = FixedAsset.query
            
            # Track applied filters
            applied_filters = {}
            
            # Apply filters if provided
            if category_id:
                query = query.filter(FixedAsset.category_id == category_id)
                applied_filters['Category ID'] = category_id
            
            if subcategory:
                query = query.join(Category).filter(Category.subcategory.ilike(f"%{subcategory}%"))
                applied_filters['Subcategory'] = subcategory
            
            # Order by ID for consistent output
            query = query.order_by(FixedAsset.id.asc())
            
            # Execute query
            assets = query.all()
            
            # Create Excel file
            return self._create_excel_export(assets, applied_filters)
            
        except Exception as e:
            print(f"Excel export error: {str(e)}")
            return create_error_response(f"Excel export failed: {str(e)}", 500)

    def _create_excel_export(self, assets, applied_filters):
        """Create Excel file with asset data and filter information"""
        try:
            # Create BytesIO buffer
            output = BytesIO()
            
            # Create workbook directly
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Assets Export"
            
            # Set column widths - removed created_at and updated_at columns
            worksheet.column_dimensions['A'].width = 10   # ID
            worksheet.column_dimensions['B'].width = 25   # Name AR
            worksheet.column_dimensions['C'].width = 25   # Name EN
            worksheet.column_dimensions['D'].width = 15   # Product Code
            worksheet.column_dimensions['E'].width = 12   # Quantity
            worksheet.column_dimensions['F'].width = 20   # Category
            worksheet.column_dimensions['G'].width = 20   # Subcategory
            worksheet.column_dimensions['H'].width = 12   # Is Active
            
            current_row = 1
            
            # Add title
            worksheet.merge_cells(f'A{current_row}:H{current_row}')  # Changed from J to H
            title_cell = worksheet[f'A{current_row}']
            title_cell.value = "Assets Export"
            title_cell.font = openpyxl.styles.Font(size=16, bold=True)
            title_cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            current_row += 2
            
            # Add generation info
            worksheet[f'A{current_row}'] = "Generated At:"
            worksheet[f'B{current_row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            current_row += 1
            
            # Add filters section (only if filters are applied)
            if applied_filters:
                worksheet[f'A{current_row}'] = "Applied Filters:"
                worksheet[f'A{current_row}'].font = openpyxl.styles.Font(bold=True)
                current_row += 1
                
                for filter_name, filter_value in applied_filters.items():
                    worksheet[f'A{current_row}'] = f"{filter_name}:"
                    worksheet[f'B{current_row}'] = filter_value
                    current_row += 1
                
                current_row += 1  # Extra spacing
            
            # Add total count
            worksheet[f'A{current_row}'] = "Total Assets:"
            worksheet[f'B{current_row}'] = len(assets)
            worksheet[f'A{current_row}'].font = openpyxl.styles.Font(bold=True)
            current_row += 2
            
            # Add data table headers - removed created_at and updated_at
            headers = [
                'ID', 'Name (Arabic)', 'Name (English)', 'Product Code', 
                'Quantity', 'Category', 'Subcategory', 'Active'
            ]
            
            header_row = current_row
            for col_num, header in enumerate(headers, 1):
                cell = worksheet.cell(row=header_row, column=col_num)
                cell.value = header
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(
                    start_color="CCCCCC", 
                    end_color="CCCCCC", 
                    fill_type="solid"
                )
            
            current_row += 1
            
            # Add asset data
            for asset in assets:
                # Get category info by querying the Category table directly
                category_name = ''
                subcategory_name = ''
                
                if asset.category_id:
                    category = db.session.get(Category, asset.category_id)
                    if category:
                        category_name = category.category or ''
                        subcategory_name = category.subcategory or ''
                
                # Removed created_at and updated_at from row_data
                row_data = [
                    asset.id,
                    asset.name_ar or '',
                    asset.name_en or '',
                    asset.product_code or '',
                    asset.quantity or 0,
                    category_name,
                    subcategory_name,
                    'Yes' if asset.is_active else 'No'
                ]
                
                for col_num, value in enumerate(row_data, 1):
                    cell = worksheet.cell(row=current_row, column=col_num)
                    cell.value = value
                
                current_row += 1
            
            # If no assets found, add a message
            if not assets:
                worksheet[f'A{current_row}'] = "No assets found for the specified filters."
                worksheet[f'A{current_row}'].font = openpyxl.styles.Font(italic=True, color="FF0000")
            
            # Save workbook to BytesIO
            workbook.save(output)
            output.seek(0)
            
            # Generate filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            if applied_filters:
                filter_suffix = "_filtered"
            else:
                filter_suffix = "_all"
            filename = f"assets_export{filter_suffix}_{timestamp}.xlsx"
            
            return send_file(
                output,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
        except Exception as e:
            print(f"Error creating Excel file: {str(e)}")
            return create_error_response(f"Failed to create Excel file: {str(e)}", 500)

